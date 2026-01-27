from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


TRUE_VALUES = {"true", "1", "yes", "y", "√", "check", "checked", "x"}
FALSE_VALUES = {"false", "0", "no", "n", ""}


@dataclass
class ParsedSheet:
    headers: list[str]
    header_rows: int
    rows: list[tuple[int, list[str]]]


def normalize_string(value: object) -> str:
    if value is None:
        return ""
    value_str = str(value).strip()
    value_str = re.sub(r"\s+", " ", value_str)
    return value_str


def normalize_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = normalize_string(value).lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return False


def normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_string(value).lower())


def parse_decimal(value: object) -> str:
    if value is None:
        return ""
    text = normalize_string(value)
    if not text:
        return ""
    text = text.replace(",", "")
    return text


def build_merged_value_map(sheet) -> dict[tuple[int, int], object]:
    merged_values: dict[tuple[int, int], object] = {}
    for merged_range in sheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        value = sheet.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = value
    return merged_values


def detect_header_rows(sheet, max_header_rows: int, resolver) -> int:
    best_rows = 1
    best_matches = -1
    merged_values = build_merged_value_map(sheet)
    max_col = sheet.max_column

    for candidate_rows in range(1, max_header_rows + 1):
        headers = []
        matches = 0
        for col in range(1, max_col + 1):
            parts = []
            for row in range(1, candidate_rows + 1):
                value = merged_values.get((row, col), sheet.cell(row=row, column=col).value)
                if value is None:
                    continue
                value_str = normalize_string(value)
                if value_str:
                    parts.append(value_str)
            header = " ".join(parts).strip()
            headers.append(header)
            if header and resolver(normalize_header(header)):
                matches += 1
        if matches > best_matches:
            best_matches = matches
            best_rows = candidate_rows
    return best_rows


def parse_excel(
    excel_path: Path,
    max_header_rows: int,
    resolver,
    sheet_name: str | None = None,
) -> ParsedSheet:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    merged_values = build_merged_value_map(sheet)
    header_rows = detect_header_rows(sheet, max_header_rows, resolver)
    max_col = sheet.max_column

    headers: list[str] = []
    for col in range(1, max_col + 1):
        parts = []
        for row in range(1, header_rows + 1):
            value = merged_values.get((row, col), sheet.cell(row=row, column=col).value)
            if value is None:
                continue
            value_str = normalize_string(value)
            if value_str:
                parts.append(value_str)
        headers.append(" ".join(parts).strip())

    rows: list[tuple[int, list[str]]] = []
    for row_idx in range(header_rows + 1, sheet.max_row + 1):
        values = []
        empty = True
        for col in range(1, max_col + 1):
            value = merged_values.get((row_idx, col), sheet.cell(row=row_idx, column=col).value)
            if value is not None and normalize_string(value):
                empty = False
            values.append(value)
        if empty:
            continue
        rows.append((row_idx, values))

    return ParsedSheet(headers=headers, header_rows=header_rows, rows=rows)


def write_csv(path: Path, columns: Iterable[str], rows: Iterable[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_bad_rows(path: Path, bad_rows: Iterable[dict[str, object]]) -> None:
    write_csv(path, ["row_number", "reason"], bad_rows)
