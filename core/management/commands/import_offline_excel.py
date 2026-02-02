from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import models, transaction

from openpyxl import load_workbook

from grms import models as grms_models
from traffic import models as traffic_models

try:
    from django.contrib.gis.geos import GEOSGeometry
except Exception:  # pragma: no cover - optional GIS dependency
    GEOSGeometry = None


@dataclass
class ImportStats:
    created: int = 0
    updated: int = 0
    errors: int = 0

    def add(self, field: str, amount: int = 1) -> None:
        setattr(self, field, getattr(self, field) + amount)


IMPORT_ORDER = [
    ("grms.Road", grms_models.Road),
    ("grms.RoadSection", grms_models.RoadSection),
    ("grms.RoadSegment", grms_models.RoadSegment),
    ("grms.RoadSocioEconomic", grms_models.RoadSocioEconomic),
    ("grms.StructureInventory", grms_models.StructureInventory),
    ("grms.BridgeDetail", grms_models.BridgeDetail),
    ("grms.CulvertDetail", grms_models.CulvertDetail),
    ("grms.RoadConditionSurvey", grms_models.RoadConditionSurvey),
    ("grms.StructureConditionSurvey", grms_models.StructureConditionSurvey),
    ("traffic.TrafficSurvey", traffic_models.TrafficSurvey),
    ("traffic.TrafficCountRecord", traffic_models.TrafficCountRecord),
]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _row_is_empty(values: tuple[Any, ...]) -> bool:
    return all(_is_empty(value) for value in values)


def _parse_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "yes", "y", "1"}:
        return True
    if text in {"false", "no", "n", "0"}:
        return False
    return None


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_time(value: Any) -> time | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return None


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _normalize_pk(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _normalize_excel_id(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _convert_field_value(field: models.Field, value: Any) -> Any:
    if _is_empty(value):
        return None
    if isinstance(field, models.BooleanField):
        parsed = _parse_bool(value)
        return parsed if parsed is not None else bool(value)
    if isinstance(field, models.DateField) and not isinstance(field, models.DateTimeField):
        return _parse_date(value)
    if isinstance(field, models.DateTimeField):
        if isinstance(value, datetime):
            return value
        parsed = _parse_date(value)
        if parsed:
            return datetime.combine(parsed, time.min)
        return None
    if isinstance(field, models.TimeField):
        return _parse_time(value)
    if isinstance(field, models.DecimalField):
        return _parse_decimal(value)
    if isinstance(field, models.IntegerField):
        if isinstance(value, int):
            return value
        if isinstance(value, float) and value.is_integer():
            return int(value)
        try:
            return int(str(value).strip())
        except Exception:
            return value
    return value


def _normalize_geom_wkt(value: str, default_srid: int) -> str:
    text = value.strip()
    if not text:
        return ""
    if text.upper().startswith("SRID="):
        return text
    return f"SRID={default_srid};{text}"


def _convert_geometry(value: Any, *, default_srid: int | None = None) -> Any:
    if _is_empty(value):
        return None
    if not isinstance(value, str):
        return value
    if GEOSGeometry is None:
        return None
    try:
        text = _normalize_geom_wkt(value, default_srid) if default_srid else value
        return GEOSGeometry(text)
    except Exception:
        return None


def _resolve_fk_value(
    model: type[models.Model],
    field_name: str,
    lookup_path: str,
    value: Any,
    survey_id_map: dict[str, int],
    current_data: dict[str, Any] | None,
    strict: bool,
    row_label: str,
) -> models.Model | None:
    if _is_empty(value):
        return None

    field = model._meta.get_field(field_name)
    if not field.is_relation:
        raise CommandError(f"{row_label}: {field_name} is not a relation field.")

    related_model = field.remote_field.model
    extra_filters: dict[str, Any] = {}
    if related_model is grms_models.ConditionFactorLookup:
        factor_type_map = {
            "drainage_left": grms_models.ConditionFactorLookup.FactorType.DRAINAGE,
            "drainage_right": grms_models.ConditionFactorLookup.FactorType.DRAINAGE,
            "shoulder_left": grms_models.ConditionFactorLookup.FactorType.SHOULDER,
            "shoulder_right": grms_models.ConditionFactorLookup.FactorType.SHOULDER,
            "surface_condition": grms_models.ConditionFactorLookup.FactorType.SURFACE,
        }
        factor_type = factor_type_map.get(field_name)
        if factor_type:
            extra_filters["factor_type"] = factor_type

    def _get_related(**kwargs: Any):
        try:
            return related_model.objects.get(**{**extra_filters, **kwargs})
        except related_model.DoesNotExist:
            if extra_filters:
                return related_model.objects.get(**kwargs)
            raise

    if (
        model is traffic_models.TrafficCountRecord
        and field_name == "traffic_survey"
        and lookup_path == "id"
    ):
        excel_id = _normalize_excel_id(value)
        mapped_id = survey_id_map.get(excel_id)
        if mapped_id is None:
            message = f"{row_label}: No TrafficSurvey mapping for excel id {excel_id}."
            if strict:
                raise CommandError(message)
            raise ValueError(message)
        return related_model.objects.get(pk=mapped_id)

    if lookup_path == "id":
        pk = _normalize_pk(value)
        if pk is not None:
            return _get_related(pk=pk)

        normalized_text = str(value).strip()
        if not normalized_text:
            message = f"{row_label}: Invalid id value for {field_name}: {value}."
            if strict:
                raise CommandError(message)
            raise ValueError(message)

        if hasattr(related_model, "description"):
            try:
                return _get_related(description__iexact=normalized_text)
            except related_model.DoesNotExist:
                pass
            try:
                return _get_related(description__istartswith=normalized_text)
            except related_model.DoesNotExist:
                pass
        if related_model is grms_models.ConditionFactorLookup:
            rating_map = {
                "good": 1,
                "fair": 2,
                "poor": 3,
                "bad": 4,
            }
            rating = rating_map.get(normalized_text.lower())
            if rating is not None:
                return _get_related(rating=rating)
        if hasattr(related_model, "name"):
            try:
                return related_model.objects.get(
                    **{**extra_filters, "name__iexact": normalized_text}
                )
            except related_model.DoesNotExist:
                pass

        message = f"{row_label}: Invalid id value for {field_name}: {value}."
        if strict:
            raise CommandError(message)
        raise ValueError(message)

    if lookup_path == "name":
        normalized_name = str(value).strip()
        if not normalized_name:
            message = f"{row_label}: Empty name value for {field_name}."
            if strict:
                raise CommandError(message)
            raise ValueError(message)
        if related_model is grms_models.AdminZone:
            obj, _ = related_model.objects.get_or_create(name=normalized_name)
            return obj
        if related_model is grms_models.AdminWoreda:
            zone = (current_data or {}).get("admin_zone")
            if zone is None:
                road = (current_data or {}).get("road")
                zone = getattr(road, "admin_zone", None)
            if zone is None:
                message = (
                    f"{row_label}: AdminWoreda name '{normalized_name}' requires admin_zone__name "
                    "in the same row or a road with admin_zone."
                )
                if strict:
                    raise CommandError(message)
                raise ValueError(message)
            obj, _ = related_model.objects.get_or_create(name=normalized_name, zone=zone)
            return obj

    lookup_value = value
    if isinstance(value, float) and value.is_integer():
        lookup_value = int(value)

    if related_model is grms_models.ConditionFactorLookup and lookup_path == "description":
        normalized_text = str(lookup_value).strip()
        try:
            return _get_related(description__iexact=normalized_text)
        except related_model.DoesNotExist:
            pass
        try:
            return _get_related(description__istartswith=normalized_text)
        except related_model.DoesNotExist:
            pass
        rating_map = {
            "good": 1,
            "fair": 2,
            "poor": 3,
            "bad": 4,
        }
        rating = rating_map.get(normalized_text.lower())
        if rating is not None:
            return _get_related(rating=rating)
        raise

    if extra_filters:
        return _get_related(**{lookup_path: lookup_value})
    return related_model.objects.get(**{lookup_path: lookup_value})


def _unique_lookup(model: type[models.Model], data: dict[str, Any], row_label: str) -> dict[str, Any]:
    if model is grms_models.Road:
        road_identifier = data.get("road_identifier")
        if not road_identifier:
            raise ValueError(f"{row_label}: Road requires road_identifier for upsert.")
        return {"road_identifier": road_identifier}
    if model is grms_models.RoadSection:
        name = data.get("name")
        road = data.get("road")
        if not name:
            raise ValueError(f"{row_label}: RoadSection requires name for upsert.")
        if road:
            return {"road": road, "name": name}
        return {"name": name}
    if model is grms_models.RoadSegment:
        segment_identifier = data.get("segment_identifier")
        if not segment_identifier:
            raise ValueError(f"{row_label}: RoadSegment requires segment_identifier for upsert.")
        return {"segment_identifier": segment_identifier}
    if model is grms_models.RoadSocioEconomic:
        road = data.get("road")
        if not road:
            raise ValueError(f"{row_label}: RoadSocioEconomic requires road for upsert.")
        return {"road": road}
    if model is grms_models.StructureInventory:
        structure_name = data.get("structure_name")
        if not structure_name:
            raise ValueError(f"{row_label}: StructureInventory requires structure_name for upsert.")
        return {"structure_name": structure_name}
    if model in {grms_models.BridgeDetail, grms_models.CulvertDetail}:
        structure = data.get("structure")
        if not structure:
            raise ValueError(f"{row_label}: {model.__name__} requires structure for upsert.")
        return {"structure": structure}
    if model is grms_models.RoadConditionSurvey:
        road_segment = data.get("road_segment")
        inspection_date = data.get("inspection_date")
        if not road_segment or not inspection_date:
            raise ValueError(f"{row_label}: RoadConditionSurvey requires road_segment and inspection_date.")
        return {"road_segment": road_segment, "inspection_date": inspection_date}
    if model is grms_models.StructureConditionSurvey:
        structure = data.get("structure")
        inspection_date = data.get("inspection_date")
        if not structure or not inspection_date:
            raise ValueError(
                f"{row_label}: StructureConditionSurvey requires structure and inspection_date."
            )
        return {"structure": structure, "inspection_date": inspection_date}
    if model is traffic_models.TrafficSurvey:
        road = data.get("road")
        survey_year = data.get("survey_year")
        cycle_number = data.get("cycle_number")
        if not road or survey_year is None or cycle_number is None:
            raise ValueError(f"{row_label}: TrafficSurvey requires road, survey_year, cycle_number.")
        return {"road": road, "survey_year": survey_year, "cycle_number": cycle_number}
    if model is traffic_models.TrafficCountRecord:
        traffic_survey = data.get("traffic_survey")
        count_date = data.get("count_date")
        if not traffic_survey or not count_date:
            raise ValueError(
                f"{row_label}: TrafficCountRecord requires traffic_survey and count_date."
            )
        return {
            "traffic_survey": traffic_survey,
            "count_date": count_date,
            "time_block_from": data.get("time_block_from"),
            "time_block_to": data.get("time_block_to"),
        }
    raise ValueError(f"{row_label}: No unique lookup configured for {model.__name__}.")


def _normalize_segment_chainage(data: dict[str, Any]) -> None:
    section = data.get("section")
    if not section:
        return
    section_length = getattr(section, "length_km", None)
    if section_length in (None, 0):
        return

    start_km = data.get("station_from_km")
    end_km = data.get("station_to_km")
    if start_km is None or end_km is None:
        return

    if start_km <= section_length and end_km <= section_length:
        return

    offset = getattr(section, "start_chainage_km", None) or Decimal("0")
    new_start = start_km - offset
    new_end = end_km - offset

    if new_start < 0 or new_end < 0:
        return
    if new_end > section_length:
        return

    data["station_from_km"] = new_start
    data["station_to_km"] = new_end


class Command(BaseCommand):
    help = "Import offline Excel data into GRMS models with idempotent upserts."

    def add_arguments(self, parser):
        parser.add_argument("path", type=Path, help="Path to the offline Excel workbook")
        parser.add_argument("--strict", action="store_true", help="Abort on first error")
        parser.add_argument("--dry-run", action="store_true", help="Validate without saving")

    def handle(self, *args, **options):
        path: Path = options["path"]
        strict: bool = options["strict"]
        dry_run: bool = options["dry_run"]

        if not path.exists():
            raise CommandError(f"Excel file not found: {path}")

        workbook = load_workbook(path, data_only=True)
        survey_id_map: dict[str, int] = {}
        total_stats = ImportStats()

        with transaction.atomic():
            for sheet_name, model in IMPORT_ORDER:
                if sheet_name not in workbook.sheetnames:
                    self.stdout.write(f"Sheet {sheet_name} not found; skipping.")
                    continue
                sheet = workbook[sheet_name]
                stats = self._import_sheet(sheet, model, sheet_name, survey_id_map, strict)
                total_stats.created += stats.created
                total_stats.updated += stats.updated
                total_stats.errors += stats.errors
                self.stdout.write(
                    f"{sheet_name}: {stats.created} created / {stats.updated} updated / {stats.errors} errors"
                )

            if dry_run:
                transaction.set_rollback(True)

        self.stdout.write(
            "Totals: "
            f"{total_stats.created} created / {total_stats.updated} updated / {total_stats.errors} errors"
        )

    def _import_sheet(
        self,
        sheet,
        model: type[models.Model],
        sheet_name: str,
        survey_id_map: dict[str, int],
        strict: bool,
    ) -> ImportStats:
        stats = ImportStats()
        rows = sheet.iter_rows(values_only=True)
        try:
            headers_row = next(rows)
        except StopIteration:
            return stats

        headers = [_normalize_header(value) for value in headers_row]
        header_keys = [header.lower() for header in headers]
        field_map = {field.name.lower(): field for field in model._meta.fields}

        for row_index, row in enumerate(rows, start=2):
            if _row_is_empty(row):
                continue
            row_label = f"{sheet_name} row {row_index}"
            excel_id: Any = None
            data: dict[str, Any] = {}

            try:
                for header, header_key, value in zip(headers, header_keys, row):
                    if not header:
                        continue
                    if _is_empty(value):
                        continue
                    if header_key == "id":
                        excel_id = value
                        continue

                    if "__" in header:
                        parts = [part.strip() for part in header.split("__") if part.strip()]
                        if len(parts) < 2:
                            continue
                        field_name = parts[0].lower()
                        lookup_path = "__".join(part.lower() for part in parts[1:])
                        field = field_map.get(field_name)
                        if not field or not field.is_relation:
                            raise ValueError(f"{row_label}: {header} is not a relation field.")
                        resolved = _resolve_fk_value(
                            model,
                            field.name,
                            lookup_path,
                            value,
                            survey_id_map,
                            data,
                            strict,
                            row_label,
                        )
                        data[field.name] = resolved
                        continue

                    field = field_map.get(header_key)
                    if not field:
                        continue
                    if field.primary_key:
                        continue
                    if field.is_relation:
                        resolved = _resolve_fk_value(
                            model,
                            field.name,
                            "id",
                            value,
                            survey_id_map,
                            data,
                            strict,
                            row_label,
                        )
                        data[field.name] = resolved
                        continue
                    if field.__class__.__name__ in {
                        "PointField",
                        "LineStringField",
                        "DjangoPointField",
                        "DjangoLineStringField",
                    }:
                        default_srid = 32637 if model is grms_models.Road and field.name == "geometry" else None
                        geometry = _convert_geometry(value, default_srid=default_srid)
                        if geometry is not None:
                            data[field.name] = geometry
                        continue
                    data[field.name] = _convert_field_value(field, value)

                if model is grms_models.RoadSegment:
                    _normalize_segment_chainage(data)

                lookup = _unique_lookup(model, data, row_label)
                defaults = {key: value for key, value in data.items() if key not in lookup}
                obj = model.objects.filter(**lookup).first()
                if obj:
                    for key, value in defaults.items():
                        setattr(obj, key, value)
                    created = False
                else:
                    obj = model(**{**lookup, **defaults})
                    created = True

                obj.full_clean()
                obj.save()
                if created:
                    stats.add("created")
                else:
                    stats.add("updated")

                if model is traffic_models.TrafficSurvey:
                    excel_key = _normalize_excel_id(excel_id) or str(row_index - 1)
                    survey_id_map[excel_key] = obj.pk
            except Exception as exc:
                stats.add("errors")
                message = f"{row_label}: {exc}"
                if strict:
                    raise CommandError(message) from exc
                self.stderr.write(message)

        return stats
