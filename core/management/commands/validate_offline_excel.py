from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import models

from openpyxl import load_workbook

from grms import models as grms_models
from traffic import models as traffic_models

try:
    from django.contrib.gis.geos import GEOSGeometry
except Exception:  # pragma: no cover - optional GIS dependency
    GEOSGeometry = None

try:
    from grms.utils import geos_length_km
except Exception:  # pragma: no cover - optional GIS dependency
    geos_length_km = None


@dataclass
class ValidationResult:
    errors: list[str]

    def add(self, message: str) -> None:
        self.errors.append(message)


MODEL_SHEETS = [
    ("grms.Road", grms_models.Road),
    ("grms.RoadSection", grms_models.RoadSection),
    ("grms.RoadSegment", grms_models.RoadSegment),
    ("grms.RoadSocioEconomic", grms_models.RoadSocioEconomic),
    ("grms.StructureInventory", grms_models.StructureInventory),
    ("grms.BridgeDetail", grms_models.BridgeDetail),
    ("grms.CulvertDetail", grms_models.CulvertDetail),
    ("grms.RoadConditionSurvey", grms_models.RoadConditionSurvey),
    ("grms.StructureConditionSurvey", grms_models.StructureConditionSurvey),
    ("traffic.TrafficSurvey", traffic_models.TrafficSurvey),
    ("traffic.TrafficCountRecord", traffic_models.TrafficCountRecord),
]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except Exception:
        return None


def _resolve_road_length_km(roads_by_id: dict[str, dict[str, Any]], sections_by_road: dict[str, list[dict[str, Any]]]) -> dict[str, float | None]:
    lengths: dict[str, float | None] = {}
    for road_id, road in roads_by_id.items():
        length = _parse_decimal(road.get("total_length_km"))
        if length and length > 0:
            lengths[road_id] = float(length)
            continue
        sec_end = [
            _parse_decimal(sec.get("end_chainage_km"))
            for sec in sections_by_road.get(road_id, [])
        ]
        sec_end = [val for val in sec_end if val is not None]
        if sec_end:
            lengths[road_id] = float(max(sec_end))
            continue
        geom = road.get("geometry")
        if geom and GEOSGeometry and geos_length_km:
            try:
                g = GEOSGeometry(str(geom))
                lengths[road_id] = float(geos_length_km(g))
                continue
            except Exception:
                pass
        lengths[road_id] = None
    return lengths


class Command(BaseCommand):
    help = "Validate offline Excel data before import."

    def add_arguments(self, parser):
        parser.add_argument("path", type=Path, help="Path to the offline Excel workbook")

    def handle(self, *args, **options):
        path: Path = options["path"]
        if not path.exists():
            raise CommandError(f"Excel file not found: {path}")

        workbook = load_workbook(path, data_only=True)
        result = ValidationResult(errors=[])

        sheets = set(workbook.sheetnames)

        sheet_headers: dict[str, list[str]] = {}
        sheet_rows: dict[str, list[tuple[Any, ...]]] = {}
        for sheet_name, _ in MODEL_SHEETS:
            if sheet_name not in sheets:
                continue
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [_normalize_header(value) for value in rows[0]]
            sheet_headers[sheet_name] = headers
            sheet_rows[sheet_name] = rows[1:]

        self._validate_required_fields(result, sheet_headers, sheet_rows)
        self._validate_sections(result, sheet_headers, sheet_rows)
        self._validate_segments(result, sheet_headers, sheet_rows)
        self._validate_structures(result, sheet_headers, sheet_rows)

        if result.errors:
            self.stdout.write(f"Found {len(result.errors)} validation errors.")
            for message in result.errors[:50]:
                self.stdout.write(message)
            if len(result.errors) > 50:
                self.stdout.write("...more errors omitted.")
            raise CommandError("Offline Excel validation failed.")

        self.stdout.write("Validation passed with no errors.")

    def _validate_required_fields(self, result, sheet_headers, sheet_rows):
        for sheet_name, model in MODEL_SHEETS:
            if sheet_name not in sheet_headers:
                continue
            headers = [h.lower() for h in sheet_headers[sheet_name]]
            header_set = set(headers)

            required_fields = []
            for field in model._meta.fields:
                if field.primary_key or field.auto_created:
                    continue
                if field.null or field.blank:
                    continue
                if field.has_default():
                    continue
                required_fields.append(field)

            for row_index, row in enumerate(sheet_rows.get(sheet_name, []), start=2):
                if all(_is_empty(value) for value in row):
                    continue
                for field in required_fields:
                    field_key = field.name.lower()
                    possible_headers = [field_key]
                    possible_headers.extend([h for h in header_set if h.startswith(f"{field_key}__")])
                    if not any(h in header_set for h in possible_headers):
                        result.add(f"{sheet_name} row {row_index}: Missing column for required field {field.name}.")
                        continue
                    value = None
                    for header in possible_headers:
                        if header in header_set:
                            idx = headers.index(header)
                            value = row[idx] if idx < len(row) else None
                            if not _is_empty(value):
                                break
                    if _is_empty(value):
                        result.add(f"{sheet_name} row {row_index}: {field.name} is required.")

    def _validate_sections(self, result, sheet_headers, sheet_rows):
        sheet_name = "grms.RoadSection"
        if sheet_name not in sheet_headers:
            return
        headers = [h.lower() for h in sheet_headers[sheet_name]]
        rows = sheet_rows.get(sheet_name, [])
        by_road: dict[str, list[tuple[Decimal, Decimal, int]]] = defaultdict(list)

        for row_index, row in enumerate(rows, start=2):
            if all(_is_empty(value) for value in row):
                continue
            road = row[headers.index("road__road_identifier")] if "road__road_identifier" in headers else None
            start = row[headers.index("start_chainage_km")] if "start_chainage_km" in headers else None
            end = row[headers.index("end_chainage_km")] if "end_chainage_km" in headers else None
            if _is_empty(road) or _is_empty(start) or _is_empty(end):
                continue
            start_d = _parse_decimal(start)
            end_d = _parse_decimal(end)
            if start_d is None or end_d is None:
                continue
            by_road[str(road)].append((start_d, end_d, row_index))

        for road_id, entries in by_road.items():
            entries.sort(key=lambda item: (item[0], item[1]))
            prev_end = None
            for start, end, row_index in entries:
                if prev_end is not None and start < prev_end:
                    result.add(
                        f"{sheet_name} row {row_index}: Section overlaps previous section on road {road_id}."
                    )
                prev_end = max(prev_end, end) if prev_end is not None else end

    def _validate_segments(self, result, sheet_headers, sheet_rows):
        sheet_name = "grms.RoadSegment"
        if sheet_name not in sheet_headers:
            return
        headers = [h.lower() for h in sheet_headers[sheet_name]]
        rows = sheet_rows.get(sheet_name, [])
        by_section: dict[str, list[tuple[Decimal, Decimal, int]]] = defaultdict(list)

        for row_index, row in enumerate(rows, start=2):
            if all(_is_empty(value) for value in row):
                continue
            section = row[headers.index("section__name")] if "section__name" in headers else None
            start = row[headers.index("station_from_km")] if "station_from_km" in headers else None
            end = row[headers.index("station_to_km")] if "station_to_km" in headers else None
            if _is_empty(section) or _is_empty(start) or _is_empty(end):
                continue
            start_d = _parse_decimal(start)
            end_d = _parse_decimal(end)
            if start_d is None or end_d is None:
                continue
            if end_d <= start_d:
                result.add(f"{sheet_name} row {row_index}: station_to_km must exceed station_from_km.")
                continue
            by_section[str(section)].append((start_d, end_d, row_index))

        for section_name, entries in by_section.items():
            entries.sort(key=lambda item: (item[0], item[1]))
            prev_end = None
            for start, end, row_index in entries:
                if prev_end is not None and start < prev_end:
                    result.add(
                        f"{sheet_name} row {row_index}: Segment overlaps previous segment in section {section_name}."
                    )
                prev_end = max(prev_end, end) if prev_end is not None else end

    def _validate_structures(self, result, sheet_headers, sheet_rows):
        sheet_name = "grms.StructureInventory"
        if sheet_name not in sheet_headers:
            return
        headers = [h.lower() for h in sheet_headers[sheet_name]]
        rows = sheet_rows.get(sheet_name, [])

        road_sheet = "grms.Road"
        section_sheet = "grms.RoadSection"
        roads = {}
        sections_by_road: dict[str, list[dict[str, Any]]] = defaultdict(list)

        if road_sheet in sheet_headers:
            road_headers = [h.lower() for h in sheet_headers[road_sheet]]
            for row in sheet_rows.get(road_sheet, []):
                if all(_is_empty(value) for value in row):
                    continue
                road_id = row[road_headers.index("road_identifier")] if "road_identifier" in road_headers else None
                if _is_empty(road_id):
                    continue
                roads[str(road_id)] = {road_headers[i]: row[i] for i in range(len(road_headers))}

        if section_sheet in sheet_headers:
            section_headers = [h.lower() for h in sheet_headers[section_sheet]]
            for row in sheet_rows.get(section_sheet, []):
                if all(_is_empty(value) for value in row):
                    continue
                road_id = row[section_headers.index("road__road_identifier")] if "road__road_identifier" in section_headers else None
                if _is_empty(road_id):
                    continue
                sections_by_road[str(road_id)].append(
                    {section_headers[i]: row[i] for i in range(len(section_headers))}
                )

        road_lengths = _resolve_road_length_km(roads, sections_by_road)

        for row_index, row in enumerate(rows, start=2):
            if all(_is_empty(value) for value in row):
                continue
            road_id = row[headers.index("road__road_identifier")] if "road__road_identifier" in headers else None
            station = row[headers.index("station_km")] if "station_km" in headers else None
            if _is_empty(road_id) or _is_empty(station):
                continue
            station_d = _parse_decimal(station)
            if station_d is None:
                continue
            length = road_lengths.get(str(road_id))
            if length is not None and station_d > Decimal(str(length)):
                result.add(
                    f"{sheet_name} row {row_index}: station_km exceeds resolved road length for {road_id}."
                )

