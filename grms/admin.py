from __future__ import annotations

import json
from io import BytesIO
from typing import Dict, List, Sequence, Tuple

import csv
from decimal import Decimal, ROUND_HALF_UP

from django import forms
from django.contrib import admin, messages
from django.contrib.admin import AdminSite
from django.contrib.admin.widgets import AutocompleteSelect
from django.db.models import Max, Min, Sum, Q
from django.template.response import TemplateResponse
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.urls import path, reverse
from openpyxl import Workbook

from . import models
from .menu import build_menu_groups
from .admin_base import GRMSBaseAdmin
from .admin_forms import CascadeFKModelFormMixin
from .forms import RoadSegmentAdminForm
from .admin_utils import valid_autocomplete_fields, valid_list_display
from .admin_mixins import CascadeAutocompleteAdminMixin, RoadSectionCascadeAutocompleteMixin
from django.db.models import Q
from traffic.models import TrafficSurveyOverall, TrafficSurveySummary
from .gis_fields import LineStringField, PointField
from .admin_cascades import (
    CascadeRoadSectionAssetMixin,
    CascadeRoadSectionMixin,
    RoadSectionCascadeAdminMixin,
    RoadSectionSegmentCascadeAdminMixin,
    RoadSectionSegmentFilterForm,
    RoadSectionStructureCascadeAdminMixin,
)
from . import admin_geojson, admin_reports
from .services import map_services, mci_intervention, prioritization, workplan_costs
from .services.planning import road_ranking, workplans
from .utils import make_point, point_to_lat_lng, utm_to_wgs84, wgs84_to_utm
from .utils_labels import (
    fmt_km,
    furniture_label,
    road_id,
    section_id,
    segment_label,
    structure_label,
)

UTM_ZONE = 37
UTM_SRID = 32637


def _serialize_geometry(geom):
    if not geom:
        return None
    if hasattr(geom, "geojson"):
        try:
            return json.loads(geom.geojson)
        except Exception:
            pass
    if isinstance(geom, (dict, list)):
        return geom
    if isinstance(geom, str):
        try:
            return json.loads(geom)
        except Exception:
            return None
    return None


def _to_float(value):
    return float(value) if value is not None else None


def _road_map_context_url(road_id):
    from django.urls import reverse

    if not road_id:
        return ""
    return reverse("road_map_context", args=[road_id])


def _reverse_or_empty(name: str, object_id):
    from django.urls import reverse

    if not object_id:
        return ""
    return reverse(name, args=[object_id])


def _filter_structure_qs(qs, request):
    road = request.GET.get("road")
    section = request.GET.get("section")
    structure_type = request.GET.get("structure_type")

    if road and road.isdigit():
        qs = qs.filter(road_id=int(road))
    if section and section.isdigit():
        qs = qs.filter(section_id=int(section))
    if structure_type and structure_type.isdigit():
        qs = qs.filter(structure_category_id=int(structure_type))
    return qs


def _workbook_response(filename: str, workbook: Workbook) -> HttpResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def export_road_segments_to_excel(modeladmin, request, queryset):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Road segments"
    ws.append(
        [
            "Road ID",
            "Section",
            "Segment",
            "Station from (km)",
            "Station to (km)",
            "Cross section",
            "Terrain (transverse)",
            "Terrain (longitudinal)",
        ]
    )
    segments = queryset.select_related("section", "section__road")
    for segment in segments:
        ws.append(
            [
                segment.section.road.road_identifier,
                section_id(segment.section),
                segment_label(segment),
                segment.station_from_km,
                segment.station_to_km,
                segment.cross_section,
                segment.terrain_transverse,
                segment.terrain_longitudinal,
            ]
        )
    return _workbook_response("road_segments.xlsx", workbook)


export_road_segments_to_excel.short_description = "Export selected to Excel"


def export_structures_to_excel(modeladmin, request, queryset):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Structures"
    ws.append(["Road ID", "Section", "Category", "Structure", "Easting (m)", "Northing (m)"])
    structures = queryset.select_related("road", "section")
    for structure in structures:
        ws.append(
            [
                structure.road.road_identifier,
                section_id(structure.section) if structure.section_id else "",
                structure.structure_category,
                structure.structure_name or "",
                structure.easting_m,
                structure.northing_m,
            ]
        )
    return _workbook_response("structures.xlsx", workbook)


export_structures_to_excel.short_description = "Export selected to Excel"


def export_condition_surveys_to_excel(modeladmin, request, queryset):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Condition surveys"
    ws.append(["Road ID", "Section", "Segment", "Inspection date", "MCI"])
    surveys = queryset.select_related("road_segment", "road_segment__section", "road_segment__section__road")
    for survey in surveys:
        mci_value = None
        if getattr(survey, "mci_result", None):
            mci_value = survey.mci_result.mci_value
        ws.append(
            [
                survey.road_segment.section.road.road_identifier,
                section_id(survey.road_segment.section),
                segment_label(survey.road_segment),
                survey.inspection_date,
                mci_value,
            ]
        )
    return _workbook_response("condition_surveys.xlsx", workbook)


export_condition_surveys_to_excel.short_description = "Export selected to Excel"


def _overlay_map_config(road_id=None, section_id=None, current_id=None):
    return {
        "road_id": road_id,
        "section_id": section_id,
        "current_id": current_id,
        "urls": {
            "road": reverse("admin:grms-geojson-road"),
            "sections": reverse("admin:grms-geojson-sections"),
            "segments": reverse("admin:grms-geojson-segments"),
            "structures": reverse("admin:grms-geojson-structures"),
        },
    }


def _coordinates_to_wgs84(coords, srid):
    if not coords:
        return None
    try:
        x, y = coords
    except Exception:
        return None
    if x is None or y is None:
        return None

    if srid == UTM_SRID:
        try:
            lat, lng = utm_to_wgs84(float(x), float(y), zone=UTM_ZONE)
            return float(lng), float(lat)
        except Exception:
            return None

    try:
        return float(x), float(y)
    except Exception:
        return None


def _utm_point(easting, northing):
    try:
        from django.contrib.gis.geos import Point

        return Point(float(easting), float(northing), srid=UTM_SRID)
    except Exception:
        return {
            "type": "Point",
            "coordinates": [float(easting), float(northing)],
            "srid": UTM_SRID,
        }


def _point_to_wgs84(point):
    if not point:
        return None

    srid = None
    coords = None

    if isinstance(point, dict):
        srid = point.get("srid")
        coords = point.get("coordinates")
    else:
        srid = getattr(point, "srid", None)
        coords = (getattr(point, "x", None), getattr(point, "y", None))

    if hasattr(point, "transform"):
        try:
            geom_4326 = point.transform(4326, clone=True) if srid and srid != 4326 else point
            return {"lng": float(geom_4326.x), "lat": float(geom_4326.y)}
        except Exception:
            pass

    converted = _coordinates_to_wgs84(coords, srid)
    if converted:
        lng, lat = converted
        return {"lng": lng, "lat": lat}

    return point_to_lat_lng(point)


def _point_to_utm(point):
    if not point:
        return None, None

    srid = None
    coords = None

    if isinstance(point, dict):
        srid = point.get("srid")
        coords = point.get("coordinates")
    else:
        srid = getattr(point, "srid", None)
        coords = (getattr(point, "x", None), getattr(point, "y", None))

    if hasattr(point, "transform"):
        try:
            if srid and srid != UTM_SRID:
                point = point.transform(UTM_SRID, clone=True)
                srid = UTM_SRID
            if srid == UTM_SRID:
                return float(point.x), float(point.y)
        except Exception:
            pass

    if srid == UTM_SRID:
        try:
            if not coords or len(coords) < 2:
                return None, None
            return float(coords[0]), float(coords[1])
        except Exception:
            return None, None

    latlng = point_to_lat_lng(point)
    if latlng:
        try:
            return wgs84_to_utm(latlng["lat"], latlng["lng"], zone=UTM_ZONE)
        except Exception:
            return None, None

    return None, None


MenuTarget = str | Tuple[str, str]
MenuGroup = Sequence[MenuTarget] | Dict[str, Sequence[MenuTarget]]


class GRMSAdminSite(AdminSite):
    site_header = "GRMS Administration"
    site_title = "GRMS Admin"
    index_title = "Gravel Road Management System"
    index_template = "admin/index.html"
    site_url = "/"

    EXCLUDED_MODEL_NAMES: set[str] = set()

    MENU_GROUPS: Dict[str, MenuGroup] = {}

    def _get_menu_groups(self) -> Dict[str, MenuGroup]:
        """Build menu groups dynamically from the registered admin models."""
        self.MENU_GROUPS = build_menu_groups(self)
        return self.MENU_GROUPS

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "grms/options/sections/",
                self.admin_view(self.section_options_view),
                name="grms-options-sections",
            ),
            path(
                "grms/options/segments/",
                self.admin_view(self.segment_options_view),
                name="grms-options-segments",
            ),
            path(
                "grms/options/structures/",
                self.admin_view(self.structure_options_view),
                name="grms-options-structures",
            ),
            path(
                "grms/options/furniture/",
                self.admin_view(self.furniture_options_view),
                name="grms-options-furniture",
            ),
            path("grms/reports/", self.admin_view(admin_reports.reports_index_view), name="grms-reports-index"),
            path(
                "grms/reports/road-inventory.xlsx",
                self.admin_view(admin_reports.road_inventory_report_view),
                name="grms-reports-road-inventory",
            ),
            path(
                "grms/reports/structure-inventory.xlsx",
                self.admin_view(admin_reports.structure_inventory_report_view),
                name="grms-reports-structure-inventory",
            ),
            path(
                "grms/reports/condition-surveys.xlsx",
                self.admin_view(admin_reports.condition_survey_report_view),
                name="grms-reports-condition-surveys",
            ),
            path("grms/geojson/road/", self.admin_view(admin_geojson.road_geojson_view), name="grms-geojson-road"),
            path(
                "grms/geojson/sections/",
                self.admin_view(admin_geojson.sections_geojson_view),
                name="grms-geojson-sections",
            ),
            path(
                "grms/geojson/segments/",
                self.admin_view(admin_geojson.segments_geojson_view),
                name="grms-geojson-segments",
            ),
            path(
                "grms/geojson/structures/",
                self.admin_view(admin_geojson.structures_geojson_view),
                name="grms-geojson-structures",
            ),
            path(
                "grms/road-autocomplete/",
                self.admin_view(self.road_autocomplete_view),
                name="grms-road-autocomplete",
            ),
            path(
                "grms/section-autocomplete/",
                self.admin_view(self.section_autocomplete_view),
                name="grms-section-autocomplete",
            ),
        ]
        return custom + urls

    def section_options_view(self, request):
        road_id = request.GET.get("road_id")
        qs = models.RoadSection.objects.all()
        if road_id and road_id.isdigit():
            qs = qs.filter(road_id=int(road_id))
        results = [
            {"id": section.id, "text": section_id(section)}
            for section in qs.order_by("sequence_on_road", "section_number")
        ]
        return JsonResponse(results, safe=False)

    def segment_options_view(self, request):
        section_id = request.GET.get("section_id")
        qs = models.RoadSegment.objects.all()
        if section_id and section_id.isdigit():
            qs = qs.filter(section_id=int(section_id))
        results = [
            {"id": segment.id, "text": segment_label(segment)}
            for segment in qs.order_by("sequence_on_section")
        ]
        return JsonResponse(results, safe=False)

    def structure_options_view(self, request):
        road_id = request.GET.get("road_id")
        section_id = request.GET.get("section_id")
        qs = models.StructureInventory.objects.all()
        if road_id and road_id.isdigit():
            qs = qs.filter(road_id=int(road_id))
        if section_id and section_id.isdigit():
            qs = qs.filter(section_id=int(section_id))
        results = [
            {"id": structure.id, "text": structure_label(structure)}
            for structure in qs.order_by(
                "road__road_identifier",
                "section__sequence_on_road",
                "station_km",
                "start_chainage_km",
                "end_chainage_km",
                "id",
            )
        ]
        return JsonResponse(results, safe=False)

    def furniture_options_view(self, request):
        road_id = request.GET.get("road_id")
        section_id = request.GET.get("section_id")
        qs = models.FurnitureInventory.objects.select_related("section", "section__road")
        if road_id and road_id.isdigit():
            qs = qs.filter(section__road_id=int(road_id))
        if section_id and section_id.isdigit():
            qs = qs.filter(section_id=int(section_id))
        results = [
            {"id": furniture.id, "text": furniture_label(furniture)}
            for furniture in qs.order_by(
                "section__road__road_identifier",
                "section__sequence_on_road",
                "chainage_km",
                "chainage_from_km",
                "chainage_to_km",
                "id",
            )
        ]
        return JsonResponse(results, safe=False)

    def road_autocomplete_view(self, request):
        term = request.GET.get("q", "").strip()
        qs = models.Road.objects.all()
        if term:
            qs = qs.filter(
                Q(road_identifier__icontains=term)
                | Q(road_name_from__icontains=term)
                | Q(road_name_to__icontains=term)
            )
        results = [
            {"id": road.id, "label": str(road)} for road in qs.order_by("road_identifier")[:50]
        ]
        return JsonResponse({"results": results})

    def section_autocomplete_view(self, request):
        term = request.GET.get("q", "").strip()
        road_id = request.GET.get("road_id")
        qs = models.RoadSection.objects.all()
        if road_id and road_id.isdigit():
            qs = qs.filter(road_id=int(road_id))
        if term:
            qs = qs.filter(
                Q(section_number__icontains=term)
                | Q(name__icontains=term)
                | Q(road__road_identifier__icontains=term)
            )
        results = [
            {"id": section.id, "label": str(section)} for section in qs.order_by("section_number")[:50]
        ]
        return JsonResponse({"results": results})

    @staticmethod
    def _normalise(name: str) -> str:
        """
        Normalize labels to guarantee consistent matching between:
        - MENU_GROUPS entries
        - Django model object_name
        - Verbose names

        Removes spaces, underscores, and lowercases everything.
        Example:
            "RoadSection", "Road Section", "road_section" → "roadsection"
        """
        if not name:
            return ""
        return name.replace("_", "").replace(" ", "").strip().lower()

    @staticmethod
    def _parse_menu_target(target: str | Tuple[str, str]) -> Tuple[str, str]:
        """Return (lookup_label, display_label) for mixed menu definitions."""
        if isinstance(target, tuple):
            lookup_label, display_label = target
            return lookup_label, display_label

        return target, target

    @staticmethod
    def _flatten_group_models(group: MenuGroup) -> List[MenuTarget]:
        if isinstance(group, dict):
            models: List[MenuTarget] = []
            for subgroup_models in group.values():
                models.extend(subgroup_models)
            return models
        return list(group)

    def _resolve_model_by_name(self, label: str):
        normalized_label = self._normalise(label)
        for model, _admin in self._registry.items():
            model_names = {
                self._normalise(model.__name__),
                self._normalise(getattr(model._meta, "verbose_name", "")),
                self._normalise(getattr(model._meta, "model_name", "")),
            }
            if normalized_label in model_names:
                return model
        return None

    def _build_model_lookup(
        self, app_list: List[Dict[str, object]]
    ) -> Dict[str, List[Dict[str, object]]]:
        lookup: Dict[str, List[Dict[str, object]]] = {}
        for app in app_list:
            app_label = app.get("app_label")
            for model in app["models"]:
                if any(
                    self._normalise(model.get(key, "")) in self.EXCLUDED_MODEL_NAMES
                    for key in ("name", "object_name")
                ):
                    continue
                model_copy = dict(model)
                model_copy.setdefault("app_label", app_label)
                raw_name = model_copy.get("object_name") or model_copy.get("name", "")
                normalised = self._normalise(raw_name)
                lookup.setdefault(normalised, []).append(model_copy)
        return lookup

    def _build_sections(self, request, menu_groups=None) -> List[Dict[str, object]]:
        app_list = self.get_app_list(request)
        lookup = self._build_model_lookup(app_list)
        assigned: set[tuple[str | None, str]] = set()
        sections: List[Dict[str, object]] = []

        menu_groups = menu_groups or self._get_menu_groups()
        for title, models_group in menu_groups.items():
            display_title = title.replace("_", " ").strip()
            grouped_models: List[Dict[str, object]] = []
            for target in self._flatten_group_models(models_group):
                lookup_label, display_name = self._parse_menu_target(target)
                for model in lookup.get(self._normalise(lookup_label), []):
                    identifier = (model.get("app_label"), model["object_name"])
                    if identifier in assigned:
                        continue
                    model_entry = dict(model)
                    model_entry["name"] = display_name
                    grouped_models.append(model_entry)
                    assigned.add(identifier)
            if grouped_models:
                sections.append({"title": display_title, "models": grouped_models})

        leftovers: List[Dict[str, object]] = []
        other_section = next(
            (section for section in sections if section["title"] == "Other models"),
            None,
        )
        for models_group in lookup.values():
            for model in models_group:
                identifier = (model.get("app_label"), model["object_name"])
                if identifier not in assigned:
                    leftovers.append(model)
                    assigned.add(identifier)
        if leftovers:
            if other_section:
                other_section["models"].extend(leftovers)
            else:
                sections.append({"title": "Other models", "models": leftovers})
        return sections

    def each_context(self, request):
        context = super().each_context(request)
        menu_groups = self._get_menu_groups()
        context["sections"] = self._build_sections(request, menu_groups=menu_groups)
        context["MENU_GROUPS"] = menu_groups
        context["get_model_by_name"] = self._resolve_model_by_name
        return context

    def index(self, request, extra_context=None):
        app_list = self.get_app_list(request)
        base_ctx = self.each_context(request) or {}
        if not isinstance(base_ctx, dict):
            base_ctx = dict(base_ctx)

        if extra_context is None:
            extra_context = {}
        elif not isinstance(extra_context, dict):
            extra_context = dict(extra_context)

        def with_default(labels, data, default_label="No data"):
            labels = list(labels)
            data = list(data)
            if not labels:
                return json.dumps([default_label]), json.dumps([0])
            return json.dumps(labels), json.dumps(data)

        # ------------------------------------------------------------------
        # Simple KPIs
        # ------------------------------------------------------------------
        total_roads = models.Road.objects.count()
        total_sections = models.RoadSection.objects.count()
        total_segments = models.RoadSegment.objects.count()
        planned_interventions = (
            models.StructureIntervention.objects.count()
            + models.RoadSectionIntervention.objects.count()
        )
        total_road_km = (
            models.Road.objects.aggregate(km=Sum("total_length_km")).get("km")
            or 0
        )
        latest_traffic_year = (
            TrafficSurveySummary.objects.aggregate(year=Max("fiscal_year"))
            .get("year")
        )

        surface_distribution = json.dumps(
            [
                models.Road.objects.filter(surface_type="Gravel").count(),
                models.Road.objects.filter(surface_type="Paved").count(),
            ]
        )

        # ------------------------------------------------------------------
        # Traffic distribution by vehicle class
        # ------------------------------------------------------------------
        traffic_qs = (
            TrafficSurveySummary.objects.values("vehicle_class")
            .annotate(total=Sum("adt_final"))
            .order_by("vehicle_class")
        )
        traffic_labels = [
            entry.get("vehicle_class") or "Unspecified" for entry in traffic_qs
        ]
        traffic_data = [
            float(entry.get("total") or 0)
            for entry in traffic_qs
        ]
        traffic_labels, traffic_data = with_default(traffic_labels, traffic_data)

        # ------------------------------------------------------------------
        # Network condition distribution based on SegmentMCIResult
        # ------------------------------------------------------------------
        condition_counts = {"good": 0, "fair": 0, "poor": 0, "bad": 0}
        for mci in models.SegmentMCIResult.objects.values_list("mci_value", flat=True):
            value = float(mci)
            if value >= 75:
                condition_counts["good"] += 1
            elif value >= 50:
                condition_counts["fair"] += 1
            elif value >= 25:
                condition_counts["poor"] += 1
            else:
                condition_counts["bad"] += 1

        condition_distribution = json.dumps(
            [
                condition_counts["good"],
                condition_counts["fair"],
                condition_counts["poor"],
                condition_counts["bad"],
            ]
        )

        # MCI histogram bins – counts of segments/surveys per MCI range
        mci_bins = [(0, 20), (21, 40), (41, 60), (61, 80), (81, 100)]
        mci_counts = json.dumps(
            [
                models.SegmentMCIResult.objects.filter(
                    mci_value__gte=lower, mci_value__lte=upper
                ).count()
                for lower, upper in mci_bins
            ]
        )
        mci_bins_labels = json.dumps([f"{lower}-{upper}" for lower, upper in mci_bins])
        mci_bins_labels, mci_counts = with_default(
            json.loads(mci_bins_labels), json.loads(mci_counts)
        )

        # ------------------------------------------------------------------
        # Network length by zone
        # ------------------------------------------------------------------
        zone_lengths_qs = (
            models.Road.objects.values("admin_zone__name")
            .annotate(total=Sum("total_length_km"))
            .order_by("admin_zone__name")
        )
        zone_labels = json.dumps(
            [entry.get("admin_zone__name") or "Unspecified" for entry in zone_lengths_qs]
        )
        zone_lengths = json.dumps(
            [float(entry.get("total") or 0) for entry in zone_lengths_qs]
        )
        zone_labels, zone_lengths = with_default(
            json.loads(zone_labels), json.loads(zone_lengths)
        )

        # ------------------------------------------------------------------
        # Prioritization summary
        # ------------------------------------------------------------------
        priority_qs = models.PrioritizationResult.objects.exclude(priority_rank__isnull=True)
        priority_counts = json.dumps(
            [
                priority_qs.filter(priority_rank__lte=5).count(),
                priority_qs.filter(priority_rank__gt=5, priority_rank__lte=10).count(),
                priority_qs.filter(priority_rank__gt=10).count(),
            ]
        )

        context = {
            **base_ctx,
            **(extra_context or {}),
            "title": "Gravel Road Management System",
            "app_list": app_list,
            "sections": base_ctx.get("sections", []),
            "total_roads": total_roads,
            "total_sections": total_sections,
            "total_segments": total_segments,
            "planned_interventions": planned_interventions,
            "total_road_km": total_road_km,
            "latest_traffic_year": latest_traffic_year,
            "surface_distribution": surface_distribution,
            "traffic_labels": traffic_labels,
            "traffic_data": traffic_data,
            "condition_distribution": condition_distribution,
            "mci_bins": mci_bins_labels,
            "mci_counts": mci_counts,
            "zone_labels": zone_labels,
            "zone_lengths": zone_lengths,
            "priority_counts": priority_counts,
        }

        return TemplateResponse(request, "admin/index.html", context)

    def _get_model_from_label(self, label):
        """
        Resolve a label in MENU_GROUPS to a registered model class.
        For example 'Roads', 'TrafficSurvey', 'PCU lookups', etc.
        """
        for model, admin_obj in self._registry.items():
            verbose = getattr(model._meta, "verbose_name", "")
            verbose_plural = getattr(model._meta, "verbose_name_plural", "")
            if label in (verbose, verbose_plural, model.__name__):
                return model
        return None

    def app_index(self, request, app_label, extra_context=None):
        """Keep navigation consistent by redirecting per-app views to the dashboard."""
        return redirect("admin:index")


# Instantiate a single GRMSAdminSite so every generated link and template helper
# (e.g., {% url 'admin:index' %}) routes through the grouped dashboard instead
# of Django's stock admin. The custom site is mounted via project/urls.py.
grms_admin_site = GRMSAdminSite(name="admin")


class RoadAdminForm(forms.ModelForm):
    start_easting = forms.DecimalField(label="Start easting", required=False, max_digits=12, decimal_places=2)
    start_northing = forms.DecimalField(label="Start northing", required=False, max_digits=12, decimal_places=2)
    start_lat = forms.FloatField(label="Start latitude", required=False)
    start_lng = forms.FloatField(label="Start longitude", required=False)
    end_easting = forms.DecimalField(label="End easting", required=False, max_digits=12, decimal_places=2)
    end_northing = forms.DecimalField(label="End northing", required=False, max_digits=12, decimal_places=2)
    end_lat = forms.FloatField(label="End latitude", required=False)
    end_lng = forms.FloatField(label="End longitude", required=False)

    class Meta:
        model = models.Road
        exclude = ("road_start_coordinates", "road_end_coordinates")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        start = point_to_lat_lng(getattr(self.instance, "road_start_coordinates", None))
        end = point_to_lat_lng(getattr(self.instance, "road_end_coordinates", None))
        if start:
            self.fields["start_lat"].initial = start["lat"]
            self.fields["start_lng"].initial = start["lng"]
        if end:
            self.fields["end_lat"].initial = end["lat"]
            self.fields["end_lng"].initial = end["lng"]
        if getattr(self.instance, "start_easting", None) is not None:
            self.fields["start_easting"].initial = self.instance.start_easting
        if getattr(self.instance, "start_northing", None) is not None:
            self.fields["start_northing"].initial = self.instance.start_northing
        if getattr(self.instance, "end_easting", None) is not None:
            self.fields["end_easting"].initial = self.instance.end_easting
        if getattr(self.instance, "end_northing", None) is not None:
            self.fields["end_northing"].initial = self.instance.end_northing

    @staticmethod
    def _quantize_utm(value: float) -> Decimal:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def _populate_coordinates(self, prefix: str):
        lat = self.cleaned_data.get(f"{prefix}_lat")
        lng = self.cleaned_data.get(f"{prefix}_lng")
        easting = self.cleaned_data.get(f"{prefix}_easting")
        northing = self.cleaned_data.get(f"{prefix}_northing")

        has_lat = lat is not None
        has_lng = lng is not None
        has_easting = easting is not None
        has_northing = northing is not None

        latlng_complete = has_lat and has_lng
        utm_complete = has_easting and has_northing

        if latlng_complete and utm_complete:
            return {"lat": float(lat), "lng": float(lng)}

        if latlng_complete:
            try:
                easting_val, northing_val = wgs84_to_utm(float(lat), float(lng), zone=37)
            except ImportError as exc:
                raise forms.ValidationError(str(exc))
            self.cleaned_data[f"{prefix}_easting"] = self._quantize_utm(easting_val)
            self.cleaned_data[f"{prefix}_northing"] = self._quantize_utm(northing_val)
            return {"lat": float(lat), "lng": float(lng)}

        if utm_complete:
            try:
                lat_val, lon_val = utm_to_wgs84(float(easting), float(northing), zone=37)
            except ImportError as exc:
                raise forms.ValidationError(str(exc))
            self.cleaned_data[f"{prefix}_lat"] = lat_val
            self.cleaned_data[f"{prefix}_lng"] = lon_val
            return {"lat": lat_val, "lng": lon_val}

        if has_easting or has_northing:
            missing = "northing" if has_easting else "easting"
            raise forms.ValidationError({f"{prefix}_{missing}": "Provide both easting and northing or a latitude/longitude pair."})

        if has_lat or has_lng:
            missing = "lng" if has_lat else "lat"
            raise forms.ValidationError({f"{prefix}_{missing}": "Provide both latitude and longitude or a UTM easting/northing pair."})

        return None

    def _clean_point(self, prefix: str):
        lat = self.cleaned_data.get(f"{prefix}_lat")
        lng = self.cleaned_data.get(f"{prefix}_lng")
        if lat is None and lng is None:
            return None
        return make_point(lat, lng)

    def clean(self):
        cleaned = super().clean()
        self._populate_coordinates("start")
        self._populate_coordinates("end")
        start_point = self._clean_point("start")
        end_point = self._clean_point("end")
        cleaned["road_start_coordinates"] = start_point
        cleaned["road_end_coordinates"] = end_point
        return cleaned

    def save(self, commit=True):
        instance = super().save(commit=False)
        instance.road_start_coordinates = self.cleaned_data.get("road_start_coordinates")
        instance.road_end_coordinates = self.cleaned_data.get("road_end_coordinates")
        if commit:
            instance.save()
            self.save_m2m()
        return instance


class SectionScopedAdmin(GRMSBaseAdmin):
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "section":
            road_id = request.POST.get("road") or request.GET.get("road")
            if road_id:
                kwargs["queryset"] = models.RoadSection.objects.filter(road_id=road_id)
        if db_field.name in {"road_segment", "segment"}:
            section_id = request.POST.get("section") or request.GET.get("section")
            if section_id:
                kwargs["queryset"] = models.RoadSegment.objects.filter(section_id=section_id)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class RoadAdmin(SectionScopedAdmin):
    form = RoadAdminForm
    list_display = (
        "road_identifier",
        "road_name_from",
        "road_name_to",
        "admin_zone",
        "admin_woreda",
        "surface_type",
        "total_length_km",
    )
    list_filter = (
        "admin_zone",
        "admin_woreda",
        "surface_type",
        "managing_authority",
        "design_standard",
    )
    search_fields = (
        "road_identifier",
        "road_name_from",
        "road_name_to",
        "admin_woreda__name",
        "admin_zone__name",
    )
    autocomplete_fields = ("admin_zone", "admin_woreda")
    change_form_template = "admin/grms/road/change_form.html"
    fieldsets = (
        (
            "Administrative context",
            {
                "fields": (
                    "road_identifier",
                    ("road_name_from", "road_name_to"),
                    ("admin_zone", "admin_woreda"),
                    "year_of_update",
                )
            },
        ),
        (
            "Classification",
            {
                "fields": (
                    "managing_authority",
                    "design_standard",
                    "surface_type",
                )
            },
        ),
        (
            "Physical characteristics",
            {
                "fields": (
                    "total_length_km",
                    "remarks",
                )
            },
        ),
        (
            "Alignment coordinates",
            {
                "description": "Capture the start and end of the road in UTM (Zone 37N) or decimal degrees (WGS84).",
                "fields": (
                    ("start_easting", "start_northing"),
                    ("end_easting", "end_northing"),
                    ("start_lat", "start_lng"),
                    ("end_lat", "end_lng"),
                ),
            },
        ),
    )

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        extra_context = extra_context or {}
        road_id = int(object_id) if object_id and object_id.isdigit() else None
        if road_id:
            map_context_url = _reverse_or_empty("road_map_context", road_id)
        else:
            from django.urls import reverse as _reverse
            map_context_url = _reverse("road_map_context_default")
        extra_context["road_admin_config"] = {
            "road_id": road_id,
            "api": {
                "route": _reverse_or_empty("road_route", road_id),
                "geometry": _reverse_or_empty("road_geometry", road_id),
                "map_context": map_context_url,
            },
        }
        extra_context["travel_modes"] = sorted(map_services.TRAVEL_MODES)
        return super().changeform_view(request, object_id, form_url, extra_context)


class RoadSectionAdminForm(CascadeFKModelFormMixin, forms.ModelForm):
    class Meta:
        model = models.RoadSection
        exclude = (
            "section_start_coordinates",
            "section_end_coordinates",
            "start_easting",
            "start_northing",
            "end_easting",
            "end_northing",
        )


grms_admin_site.register(models.Road, RoadAdmin)


@admin.register(models.RoadGlobalCostReport, site=grms_admin_site)
class RoadGlobalCostReportAdmin(GRMSBaseAdmin):
    change_list_template = "admin/reports/global_costs.html"
    ordering = ("road_identifier",)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        return self.model.objects.none()

    def changelist_view(self, request, extra_context=None):
        rows, totals = workplan_costs.compute_global_costs_by_road()

        if request.GET.get("format") == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = (
                "attachment; filename=global_costs_by_road.csv"
            )
            writer = csv.writer(response)
            headers = [
                "Road",
                "Road length (km)",
                "RM cost",
                "PM cost",
                "Rehab cost",
                "Road bottleneck cost",
                "Structure bottleneck cost",
                "Total cost",
            ]
            writer.writerow(headers)
            for row in rows:
                writer.writerow(
                    [
                        str(row["road"]),
                        row["road_length_km"],
                        row["rm_cost"],
                        row["pm_cost"],
                        row["rehab_cost"],
                        row["road_bneck_cost"],
                        row["structure_bneck_cost"],
                        row["total_cost"],
                    ]
                )
            writer.writerow(
                [
                    "Total",
                    totals.get("road_length_km", Decimal("0")),
                    totals.get("rm_cost", Decimal("0")),
                    totals.get("pm_cost", Decimal("0")),
                    totals.get("rehab_cost", Decimal("0")),
                    totals.get("road_bneck_cost", Decimal("0")),
                    totals.get("structure_bneck_cost", Decimal("0")),
                    totals.get("total_cost", Decimal("0")),
                ]
            )
            return response

        context = {
            **self.admin_site.each_context(request),
            "title": "Global Cost of Road Works",
            "opts": self.model._meta,
            "rows": rows,
            "totals": totals,
            "bucket_labels": [
                ("rm_cost", "RM cost"),
                ("pm_cost", "PM cost"),
                ("rehab_cost", "Rehab cost"),
                ("road_bneck_cost", "Road bottleneck cost"),
                ("structure_bneck_cost", "Structure bottleneck cost"),
            ],
            "csv_export_url": f"{request.path}?format=csv",
        }
        return TemplateResponse(request, self.change_list_template, context)


@admin.register(models.SectionWorkplanReport, site=grms_admin_site)
class SectionWorkplanReportAdmin(GRMSBaseAdmin):
    change_list_template = "admin/grms/reports/section_workplan.html"

    def has_add_permission(self, request):  # pragma: no cover - report only
        return False

    def has_change_permission(self, request, obj=None):  # pragma: no cover - report only
        return False

    def has_delete_permission(self, request, obj=None):  # pragma: no cover - report only
        return False

    def get_queryset(self, request):
        return self.model.objects.none()

    def changelist_view(self, request, extra_context=None):
        fiscal_year = request.GET.get("fiscal_year")
        road_id = request.GET.get("road")
        rows = []
        totals = {}
        header_context = {}

        selected_road = None
        if fiscal_year and road_id:
            try:
                selected_road = models.Road.objects.get(pk=road_id)
                rows, totals, header_context = workplans.compute_section_workplan_rows(
                    selected_road, int(fiscal_year)
                )
            except models.Road.DoesNotExist:
                selected_road = None

        if request.GET.get("format") == "csv" and rows:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = "attachment; filename=section_workplan.csv"
            writer = csv.writer(response)
            writer.writerow(
                [
                    "Rd sec no",
                    "Start km",
                    "End km",
                    "Length km",
                    "Surface type",
                    "Surface cond",
                    "RM",
                    "PM",
                    "Rehab",
                    "Road B-neck",
                    "Struct B-neck",
                    "Year cost",
                ]
            )
            for row in rows:
                writer.writerow(
                    [
                        row.rd_sec_no,
                        row.start_km,
                        row.end_km,
                        row.length_km,
                        row.surface_type,
                        row.surface_cond,
                        row.rm_cost,
                        row.pm_cost,
                        row.rehab_cost,
                        row.road_bneck_cost,
                        row.structure_bneck_cost,
                        row.year_cost,
                    ]
                )
            writer.writerow(
                [
                    "Totals",
                    "",
                    "",
                    totals.get("length_km", Decimal("0")),
                    "",
                    "",
                    totals.get("rm_cost", Decimal("0")),
                    totals.get("pm_cost", Decimal("0")),
                    totals.get("rehab_cost", Decimal("0")),
                    totals.get("road_bneck_cost", Decimal("0")),
                    totals.get("structure_bneck_cost", Decimal("0")),
                    totals.get("year_cost", Decimal("0")),
                ]
            )
            return response

        ranked_roads = (
            models.Road.objects.filter(ranking_results__fiscal_year=fiscal_year).distinct()
            if fiscal_year
            else models.Road.objects.none()
        )

        context = {
            **self.admin_site.each_context(request),
            "title": "Section Annual Workplan (Table 25)",
            "opts": self.model._meta,
            "rows": rows,
            "totals": totals,
            "header_context": header_context,
            "fiscal_year": fiscal_year,
            "road_choices": ranked_roads,
            "selected_road_id": road_id or "",
            "csv_export_url": f"{request.path}?fiscal_year={fiscal_year}&road={road_id}&format=csv" if rows else None,
        }
        return TemplateResponse(request, self.change_list_template, context)


@admin.register(models.AnnualWorkplanReport, site=grms_admin_site)
class AnnualWorkplanReportAdmin(GRMSBaseAdmin):
    change_list_template = "admin/grms/reports/annual_workplan.html"

    def has_add_permission(self, request):  # pragma: no cover - report only
        return False

    def has_change_permission(self, request, obj=None):  # pragma: no cover - report only
        return False

    def has_delete_permission(self, request, obj=None):  # pragma: no cover - report only
        return False

    def get_queryset(self, request):
        return self.model.objects.none()

    def changelist_view(self, request, extra_context=None):
        fiscal_year = request.GET.get("fiscal_year")
        group = request.GET.get("group") or None
        budget_cap = request.GET.get("budget_cap") or None
        include_partial_last_road = request.GET.get("include_partial_last_road", "true")
        include_partial_last_road = include_partial_last_road.lower() != "false"
        rows = []
        totals = {}
        header_context = {}

        if fiscal_year:
            rows, totals, header_context = workplans.compute_annual_workplan_rows(
                int(fiscal_year),
                group=group,
                budget_cap_birr=Decimal(budget_cap) if budget_cap else None,
                include_partial_last_road=include_partial_last_road,
            )

        if request.GET.get("format") == "csv" and rows:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = "attachment; filename=annual_workplan.csv"
            writer = csv.writer(response)
            writer.writerow(
                [
                    "Road no",
                    "Road class",
                    "Length km",
                    "Rank",
                    "Funding Status",
                    "Funded %",
                    "Funded Amount",
                    "RM",
                    "PM",
                    "Rehab",
                    "Road B-neck",
                    "Struct B-neck",
                    "Year cost",
                ]
            )
            for row in rows:
                writer.writerow(
                    [
                        row["road_no"],
                        row["road_class"],
                        row["road_length_km"],
                        row["rank"],
                        row.get("funding_status", "FULL"),
                        row.get("funded_percent", Decimal("100")),
                        row.get("funded_amount", row.get("year_cost")),
                        row.get("funded_rm_cost", row["rm_cost"]),
                        row.get("funded_pm_cost", row["pm_cost"]),
                        row.get("funded_rehab_cost", row["rehab_cost"]),
                        row.get("funded_road_bneck_cost", row["road_bneck_cost"]),
                        row.get("funded_structure_bneck_cost", row["structure_bneck_cost"]),
                        row.get("year_cost"),
                    ]
                )
            writer.writerow(
                [
                    "Totals",
                    "",
                    totals.get("road_length_km", Decimal("0")),
                    "",
                    "",
                    "",
                    totals.get("year_cost", Decimal("0")),
                    totals.get("rm_cost", Decimal("0")),
                    totals.get("pm_cost", Decimal("0")),
                    totals.get("rehab_cost", Decimal("0")),
                    totals.get("road_bneck_cost", Decimal("0")),
                    totals.get("structure_bneck_cost", Decimal("0")),
                    totals.get("planned_year_cost", Decimal("0")),
                ]
            )
            return response

        context = {
            **self.admin_site.each_context(request),
            "title": "Annual Workplan (Table 26)",
            "opts": self.model._meta,
            "rows": rows,
            "totals": totals,
            "header_context": header_context,
            "fiscal_year": fiscal_year,
            "selected_group": group or "",
            "budget_cap": budget_cap or "",
            "include_partial_last_road": include_partial_last_road,
            "csv_export_url": f"{request.path}?fiscal_year={fiscal_year}&group={group}&budget_cap={budget_cap}&include_partial_last_road={str(include_partial_last_road).lower()}&format=csv" if rows else None,
        }
        return TemplateResponse(request, self.change_list_template, context)


@admin.register(models.RoadLinkTypeLookup, site=grms_admin_site)
class RoadLinkTypeLookupAdmin(GRMSBaseAdmin):
    list_display = ("name", "code", "score")
    search_fields = ("name", "code")
    fieldsets = (
        ("Road link type", {"fields": ("name", "code", "score")}),
        ("Notes", {"fields": ("notes",)}),
    )


@admin.register(models.QAStatus, site=grms_admin_site)
class QAStatusAdmin(GRMSBaseAdmin):
    list_display = ("status",)
    search_fields = ("status",)


@admin.register(models.ActivityLookup, site=grms_admin_site)
class ActivityLookupAdmin(GRMSBaseAdmin):
    list_display = ("activity_code", "activity_name", "default_unit", "is_resource_based")
    search_fields = ("activity_code", "activity_name", "notes")


class InterventionLookupAdmin(GRMSBaseAdmin):
    list_display = ("intervention_code", "name", "category", "unit_measure")
    search_fields = ("intervention_code", "name", "description")


class AnnualWorkPlanAdmin(GRMSBaseAdmin):
    list_display = ("fiscal_year", "road", "region", "woreda", "status")
    list_filter = ("fiscal_year", "status", "region")
    search_fields = ("road__road_identifier", "region", "woreda")
    autocomplete_fields = ("road",)


@admin.register(models.DistressType, site=grms_admin_site)
class DistressTypeAdmin(GRMSBaseAdmin):
    list_display = ("distress_code", "distress_name", "category")
    list_filter = ("category",)
    search_fields = ("distress_code", "distress_name", "notes")


@admin.register(models.DistressCondition, site=grms_admin_site)
class DistressConditionAdmin(GRMSBaseAdmin):
    list_display = ("distress", "severity_code", "extent_code")
    search_fields = ("distress__distress_code", "distress__distress_name", "condition_notes")
    autocomplete_fields = ("distress",)


class DistressActivityAdmin(GRMSBaseAdmin):
    _AUTO = ("condition", "activity")
    _LD = ("condition", "activity", "scale_basis")
    list_display = valid_list_display(models.DistressActivity, admin.ModelAdmin, _LD)
    search_fields = (
        "condition__distress__distress_code",
        "activity__activity_name",
        "notes",
    )
    autocomplete_fields = valid_autocomplete_fields(models.DistressActivity, _AUTO)


for model, admin_class in (
    (models.InterventionLookup, InterventionLookupAdmin),
    (models.AnnualWorkPlan, AnnualWorkPlanAdmin),
    (models.DistressActivity, DistressActivityAdmin),
):
    try:
        grms_admin_site.register(model, admin_class)
    except admin.sites.AlreadyRegistered:
        grms_admin_site.unregister(model)
        grms_admin_site.register(model, admin_class)


@admin.register(models.UnitCost, site=grms_admin_site)
class UnitCostAdmin(GRMSBaseAdmin):
    list_display = ("intervention", "region", "unit_cost", "effective_date", "expiry_date")
    search_fields = ("intervention__intervention_code", "intervention__name", "region", "notes")
    _AUTO = ("intervention",)
    autocomplete_fields = valid_autocomplete_fields(models.UnitCost, _AUTO)


@admin.register(models.AdminZone, site=grms_admin_site)
class AdminZoneAdmin(GRMSBaseAdmin):
    list_display = ("name", "region")
    search_fields = ("name", "region")
    fieldsets = (("Administrative zone", {"fields": ("name", "region")}),)


@admin.register(models.AdminWoreda, site=grms_admin_site)
class AdminWoredaAdmin(GRMSBaseAdmin):
    list_display = ("name", "zone")
    list_filter = ("zone",)
    search_fields = ("name", "zone__name")
    fieldsets = (("Woreda", {"fields": ("name", "zone")}),)


@admin.register(models.InterventionCategory, site=grms_admin_site)
class InterventionCategoryAdmin(GRMSBaseAdmin):
    list_display = ("name",)
    search_fields = ("name",)
    fieldsets = (("Intervention category", {"fields": ("name",)}),)

@admin.register(models.InterventionWorkItem, site=grms_admin_site)
class InterventionWorkItemAdmin(GRMSBaseAdmin):
    list_display = ("work_code", "description", "category", "unit", "unit_cost")
    list_filter = ("category",)
    search_fields = ("work_code", "description", "category__name")
    autocomplete_fields = ("category",)
    fieldsets = (
        ("Work item", {"fields": ("category", "work_code", "description")}),
        ("Measurement", {"fields": ("unit", "unit_cost")}),
    )


@admin.register(models.ConditionFactorLookup, site=grms_admin_site)
class ConditionFactorLookupAdmin(GRMSBaseAdmin):
    list_display = ("factor_type", "rating", "factor_value", "description")
    list_filter = ("factor_type", "rating")
    search_fields = ("description", "factor_type")


@admin.register(models.MCIWeightConfig, site=grms_admin_site)
class MCIWeightConfigAdmin(GRMSBaseAdmin):
    list_display = ("name", "effective_from", "effective_to", "is_active")
    list_filter = ("is_active",)
    search_fields = ("name",)


@admin.register(models.MCICategoryLookup, site=grms_admin_site)
class MCICategoryLookupAdmin(GRMSBaseAdmin):
    list_display = (
        "rating",
        "mci_min",
        "mci_max",
        "severity_order",
        "default_intervention",
        "is_active",
    )
    list_filter = ("is_active", "severity_order")
    search_fields = ("name", "code")


@admin.register(models.MCIRoadMaintenanceRule, site=grms_admin_site)
class MCIRoadMaintenanceRuleAdmin(GRMSBaseAdmin):
    list_display = (
        "mci_min",
        "mci_max",
        "routine",
        "periodic",
        "rehabilitation",
        "priority",
        "is_active",
    )
    list_filter = ("is_active", "routine", "periodic", "rehabilitation")
    search_fields = ("mci_min", "mci_max")
    ordering = ("priority", "mci_min")


@admin.register(models.SegmentMCIResult, site=grms_admin_site)
class SegmentMCIResultAdmin(SectionScopedAdmin):
    list_display = ("road_segment", "survey_date", "mci_value", "rating")
    list_filter = ("survey_date", "rating")
    readonly_fields = ("computed_at",)
    _AUTO = ("road_segment",)
    autocomplete_fields = valid_autocomplete_fields(models.SegmentMCIResult, _AUTO)


@admin.register(models.SegmentInterventionRecommendation, site=grms_admin_site)
class SegmentInterventionRecommendationAdmin(SectionScopedAdmin):
    list_display = ("road", "section", "segment", "mci_value", "recommended_item")
    search_fields = (
        "segment__section__road__road_identifier",
        "recommended_item__work_code",
    )
    list_select_related = ("segment", "segment__section", "segment__section__road", "recommended_item")
    list_filter = ("recommended_item__category", "recommended_item__work_code")
    autocomplete_fields = ("segment", "recommended_item")

    def road(self, obj):
        return road_id(obj.segment.section.road)

    def section(self, obj):
        return section_id(obj.segment.section)

    def segment(self, obj):
        return segment_label(obj.segment)


@admin.register(models.StructureInterventionRecommendation, site=grms_admin_site)
class StructureInterventionRecommendationAdmin(GRMSBaseAdmin):
    list_display = ("structure_desc", "condition_code", "recommended_item_display")
    search_fields = ("structure__road__road_identifier", "structure__structure_category")
    list_select_related = ("structure", "structure__road", "structure__section", "recommended_item")
    autocomplete_fields = ("structure", "recommended_item")

    def structure_desc(self, obj):
        return structure_label(obj.structure)

    structure_desc.short_description = "Structure"

    @admin.display(description="Recommended item", ordering="recommended_item__work_code")
    def recommended_item_display(self, obj):
        wi = obj.recommended_item
        return f"{wi.work_code} - {wi.description}"

@admin.register(models.RoadSection, site=grms_admin_site)
class RoadSectionAdmin(RoadSectionCascadeAutocompleteMixin, RoadSectionCascadeAdminMixin, SectionScopedAdmin):
    form = RoadSectionAdminForm
    list_display = (
        "road",
        "section_number",
        "sequence_on_road",
        "length_km",
        "surface_type",
    )
    ordering = ("road__road_identifier", "id")
    list_filter = ("admin_zone_override", "admin_woreda_override", "surface_type")
    search_fields = ("section_number", "name", "road__road_identifier")
    autocomplete_fields = ("road", "admin_zone_override", "admin_woreda_override")
    readonly_fields = ("section_number", "sequence_on_road", "length_km")
    fieldsets = (
        ("Parent road", {"fields": ("road",), "description": "Select the road this section belongs to."}),
        (
            "Section identification",
            {"fields": (("section_number", "sequence_on_road"), "name")},
        ),
        (
            "Chainage and length",
            {
                "fields": (("start_chainage_km", "end_chainage_km"), "length_km"),
                "description": "Chainages must not overlap other sections on this road.",
            },
        ),
        (
            "Physical characteristics",
            {"fields": ("surface_type", "surface_thickness_cm")},
        ),
        (
            "Administrative overrides",
            {
                "description": "Use only when the section crosses into a new zone or woreda; otherwise the parent road values apply.",
                "fields": ("admin_zone_override", "admin_woreda_override"),
            },
        ),
        ("Notes", {"fields": ("notes",)}),
    )

    def get_fieldsets(self, request, obj=None):
        """Ensure fields are not repeated across fieldsets to satisfy admin checks."""

        cleaned_fieldsets = []
        seen_fields = set()

        for name, options in super().get_fieldsets(request, obj):
            fields = options.get("fields", ())
            normalised = []

            for entry in fields:
                if isinstance(entry, (list, tuple)):
                    row = [field for field in entry if field not in seen_fields]
                    if row:
                        normalised.append(tuple(row))
                        seen_fields.update(row)
                else:
                    if entry not in seen_fields:
                        normalised.append(entry)
                        seen_fields.add(entry)

            options = dict(options)
            options["fields"] = tuple(normalised)
            cleaned_fieldsets.append((name, options))

        return tuple(cleaned_fieldsets)

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        road_id = (
            request.GET.get("forward[road]")
            or request.GET.get("forward[road_id]")
            or request.GET.get("forward[road__id__exact]")
            or request.GET.get("road")
            or request.GET.get("road_id")
            or request.GET.get("road__id__exact")
        )
        if road_id and road_id.isdigit():
            queryset = queryset.filter(road_id=int(road_id))
        return queryset, use_distinct

    def get_urls(self):
        return super().get_urls()


@admin.register(models.RoadSegment, site=grms_admin_site)
class RoadSegmentAdmin(RoadSectionCascadeAdminMixin, SectionScopedAdmin):
    form = RoadSegmentAdminForm
    list_display = (
        "road",
        "section_label",
        "segment_label",
        "station_from_km",
        "station_to_km",
        "cross_section",
    )
    search_fields = (
        "section__road__road_identifier",
        "section__road__road_name_from",
        "section__road__road_name_to",
        "section__section_number",
    )
    list_filter = ("section__road", "section", "terrain_longitudinal", "terrain_transverse")
    autocomplete_fields = ("section",)
    actions = [export_road_segments_to_excel]
    fieldsets = (
        (
            "Context",
            {
                "fields": ("road", "section"),
                "description": "Select Road first (type-to-search) to filter Sections. Then select Section (type-to-search).",
            },
        ),
        (
            "Chainage",
            {
                "fields": (("station_from_km", "station_to_km"), "carriageway_width_m"),
                "description": "Segment chainages cannot overlap other segments in this section.",
            },
        ),
        (
            "Attributes",
            {
                "fields": (
                    "cross_section",
                    "terrain_transverse",
                    "terrain_longitudinal",
                )
            },
        ),
        (
            "Roadway elements",
            {
                "fields": (
                    ("ditch_left_present", "ditch_right_present"),
                    ("shoulder_left_present", "shoulder_right_present"),
                )
            },
        ),
        ("Notes", {"fields": ("comment",)}),
    )

    @admin.display(description="Road", ordering="section__road__road_identifier")
    def road(self, obj):
        road = getattr(getattr(obj, "section", None), "road", None)
        return str(road) if road else "—"

    def section_label(self, obj):
        sec = obj.section
        return f"{sec.road.road_identifier}-S{sec.sequence_on_road}" if sec else ""

    section_label.short_description = "Section"

    def segment_label(self, obj):
        return obj.segment_identifier or obj.segment_label

    class Media:
        js = ("grms/js/roadsegment_cascade.js",)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        field = super().formfield_for_foreignkey(db_field, request, **kwargs)
        if db_field.name == "section":
            road_id = (
                request.POST.get("road")
                or request.GET.get("road_id")
                or request.GET.get("road__id__exact")
            )
            if road_id and str(road_id).isdigit():
                field.queryset = models.RoadSection.objects.filter(
                    road_id=int(road_id)
                ).order_by("sequence_on_road", "id")
            else:
                field.queryset = models.RoadSection.objects.none()
        return field

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        road_id = request.GET.get("road__id__exact")
        if road_id and road_id.isdigit():
            qs = qs.filter(section__road_id=int(road_id))
        return qs

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        section_id = request.GET.get("section_id")
        if section_id and section_id.isdigit():
            queryset = queryset.filter(section_id=int(section_id))
        return queryset, use_distinct

@admin.register(models.StructureInventory, site=grms_admin_site)
class StructureInventoryAdmin(RoadSectionCascadeAutocompleteMixin, GRMSBaseAdmin):
    class StructureInventoryAdminForm(CascadeFKModelFormMixin, CascadeRoadSectionMixin, forms.ModelForm):
        class Meta:
            model = models.StructureInventory
            fields = "__all__"

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.fields["station_km"].label = "Chainage (km)"
            self._setup_road_section()

            instance = self.instance
            if instance and getattr(instance, "location_point", None):
                easting, northing = _point_to_utm(instance.location_point)
                if easting is not None and instance.easting_m is None:
                    self.fields["easting_m"].initial = float(easting)
                if northing is not None and instance.northing_m is None:
                    self.fields["northing_m"].initial = float(northing)

        def clean(self):
            cleaned = super().clean()
            easting = cleaned.get("easting_m")
            northing = cleaned.get("northing_m")
            station_km = cleaned.get("station_km")
            geometry_type = cleaned.get("geometry_type") or getattr(
                self.instance, "geometry_type", models.StructureInventory.POINT
            )

            if (easting is None) ^ (northing is None):
                missing = "northing_m" if easting is not None else "easting_m"
                raise forms.ValidationError({missing: "Provide both Easting and Northing."})

            has_en = easting is not None and northing is not None
            if (
                geometry_type == models.StructureInventory.POINT
                and cleaned.get("location_point") is None
                and station_km is None
                and not has_en
            ):
                raise forms.ValidationError(
                    "Provide Easting & Northing, or Chainage (km), or select a Location point on the map."
                )

            return cleaned

    geometry_widget = forms.Textarea(attrs={"rows": 3})

    list_display = ("label", "structure_category", "road", "section")
    list_filter = ("structure_category", "geometry_type")
    search_fields = (
        "structure_name",
        "road__road_identifier",
        "section__section_number",
        "structure_category",
    )
    list_select_related = ("road", "section")
    readonly_fields = ("created_date", "modified_date", "derived_lat_lng")
    form = StructureInventoryAdminForm
    autocomplete_fields = ("road", "section")
    actions = [export_structures_to_excel]
    formfield_overrides = {
        PointField: {"widget": geometry_widget},
        LineStringField: {"widget": geometry_widget},
    }

    fieldsets = (
        (
            "Structure",
            {
                "fields": (
                    "road",
                    "section",
                    "structure_category",
                    "structure_name",
                )
            },
        ),
        (
            "Point location",
            {
                "classes": ("structure-point",),
                "fields": (
                    ("easting_m", "northing_m", "utm_zone"),
                    "location_point",
                    "derived_lat_lng",
                ),
                "description": "Enter UTM Easting/Northing or click the map to populate coordinates.",
            },
        ),
        (
            "Chainage",
            {
                "classes": ("collapse", "structure-point"),
                "fields": ("station_km",),
                "description": "Chainage reference for point structures.",
            },
        ),
        (
            "Line location",
            {
                "classes": ("structure-line",),
                "fields": (
                    "start_chainage_km",
                    "end_chainage_km",
                    "location_line",
                ),
            },
        ),
        ("Documentation", {"fields": ("comments", "attachments")}),
        ("Timestamps", {"fields": ("created_date", "modified_date")}),
    )

    class Media:
        js = ("grms/admin/cascade_autocomplete.js",)

    class Media:
        js = (
            "grms/js/structure-inventory-admin.js",
            "grms/admin/cascade_autocomplete.js",
        )

    def derived_lat_lng(self, obj):
        point = _point_to_wgs84(obj.location_point) if obj and obj.location_point else None
        if not point:
            return "—"
        return f"{point['lat']:.6f}, {point['lng']:.6f}"

    derived_lat_lng.short_description = "Lat/Lng"

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        extra_context = extra_context or {}
        instance = self.get_object(request, object_id)
        road_id = instance.road_id if instance else request.GET.get("road")
        section_id = instance.section_id if instance else request.GET.get("section")
        current_id = instance.id if instance else None
        extra_context["overlay_map_config"] = _overlay_map_config(
            road_id=road_id,
            section_id=section_id,
            current_id=current_id,
        )
        return super().changeform_view(request, object_id, form_url, extra_context)

    def label(self, obj):
        return structure_label(obj)

    label.short_description = "Structure"

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        road_id = request.GET.get("road_id")
        if road_id and road_id.isdigit():
            queryset = queryset.filter(road_id=int(road_id))
        section_id = request.GET.get("section_id")
        if section_id and section_id.isdigit():
            queryset = queryset.filter(section_id=int(section_id))
        return queryset, use_distinct

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "structures_geojson/",
                self.admin_site.admin_view(self.structures_geojson_view),
                name="grms_structureinventory_structures_geojson",
            ),
        ]
        return custom + urls

    def structures_geojson_view(self, request):
        road_id = request.GET.get("road_id")
        if not road_id:
            return JsonResponse({"error": "road_id is required."}, status=400)

        qs = models.StructureInventory.objects.filter(road_id=road_id).exclude(location_point__isnull=True)

        exclude_id = request.GET.get("exclude_id")
        if exclude_id and exclude_id.isdigit():
            qs = qs.exclude(id=exclude_id)

        current_id = None
        try:
            current_id = int(request.GET.get("current_id", "") or 0)
        except (TypeError, ValueError):
            current_id = None

        features = []
        for structure in qs:
            point_latlng = _point_to_wgs84(structure.location_point)
            coords = None
            if point_latlng:
                coords = (point_latlng.get("lng"), point_latlng.get("lat"))

            if not coords:
                continue

            properties = {
                "id": structure.id,
                "category": structure.structure_category,
                "label": structure_label(structure),
                "station_km": float(structure.station_km) if structure.station_km is not None else None,
                "name": structure.structure_name,
                "is_current": bool(current_id and structure.id == current_id),
            }
            features.append(
                {
                    "type": "Feature",
                    "geometry": {"type": "Point", "coordinates": [coords[0], coords[1]]},
                    "properties": properties,
                }
            )

        return JsonResponse({"type": "FeatureCollection", "features": features})


class StructureDetailOverlayMixin:
    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        extra_context = extra_context or {}
        instance = self.get_object(request, object_id)
        structure = getattr(instance, "structure", None) if instance else None
        road_id = structure.road_id if structure else request.GET.get("road")
        section_id = structure.section_id if structure else request.GET.get("section")
        current_id = structure.id if structure else None
        extra_context["overlay_map_config"] = _overlay_map_config(
            road_id=road_id,
            section_id=section_id,
            current_id=current_id,
        )
        return super().changeform_view(request, object_id, form_url, extra_context)


class StructureDetailFilterForm(CascadeRoadSectionAssetMixin, forms.ModelForm):
    structure_category: str | None = None
    asset_field_name = "structure"
    asset_model = models.StructureInventory
    asset_url = "/admin/grms/options/structures/"
    asset_placeholder = "Select a structure"

    road = forms.ModelChoiceField(
        queryset=models.Road.objects.all(),
        required=False,
        label="Road",
    )
    section = forms.ModelChoiceField(
        queryset=models.RoadSection.objects.all(),
        required=False,
        label="Section",
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        road_field = self.fields.get("road")
        if road_field is not None:
            road_field.widget = AutocompleteSelect(
                models.RoadSection._meta.get_field("road"),
                grms_admin_site,
            )
        section_field = self.fields.get("section")
        if section_field is not None:
            section_field.widget = AutocompleteSelect(
                models.StructureInventory._meta.get_field("section"),
                grms_admin_site,
            )
        instance = self.instance
        if instance and getattr(instance, "structure_id", None):
            structure = instance.structure
            self.fields["road"].initial = structure.road
            self.fields["section"].initial = structure.section
        self._setup_road_section()
        self._setup_asset()

    def _asset_queryset(self, road_id: str | None, section_id: str | None):
        qs = super()._asset_queryset(road_id, section_id)
        if self.structure_category:
            qs = qs.filter(structure_category=self.structure_category)
        return qs

    def clean(self):
        cleaned = super().clean()
        structure = cleaned.get("structure")

        if structure and self.structure_category and structure.structure_category != self.structure_category:
            raise forms.ValidationError({"structure": "Selected structure must match the required category."})
        return cleaned


class BridgeDetailForm(StructureDetailFilterForm):
    structure_category = "Bridge"

    class Meta:
        model = models.BridgeDetail
        fields = "__all__"


@admin.register(models.BridgeDetail, site=grms_admin_site)
class BridgeDetailAdmin(StructureDetailOverlayMixin, RoadSectionStructureCascadeAdminMixin, GRMSBaseAdmin):
    form = BridgeDetailForm
    structure_category_codes = ("Bridge",)
    list_display = ("structure", "bridge_type", "span_count", "has_head_walls")
    autocomplete_fields = ("structure",)
    cascade_autocomplete = {
        "structure": lambda qs, req: _filter_structure_qs(qs, req),
    }
    change_form_template = "admin/grms/structure_detail/change_form.html"
    class Media:
        js = (
            "grms/admin/cascade_autocomplete.js",
        )
    fieldsets = (
        ("Structure", {"fields": ("road", "section", "structure")}),
        (
            "Bridge details",
            {
                "fields": (
                    "bridge_type",
                    ("span_count", "width_m", "length_m"),
                    "has_head_walls",
                )
            },
        ),
    )


class CulvertDetailForm(StructureDetailFilterForm):
    structure_category = "Culvert"

    class Meta:
        model = models.CulvertDetail
        fields = "__all__"

    class Media:
        js = ("grms/js/culvert-detail-admin.js",)

    slab_box_fields = ("width_m", "span_m", "clear_height_m")
    pipe_fields = ("num_pipes", "pipe_diameter_m")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        culvert_type = self.initial.get("culvert_type") or getattr(self.instance, "culvert_type", None)
        self._apply_field_state(culvert_type)

    def _apply_field_state(self, culvert_type: str | None):
        def set_disabled(field_names, disabled: bool):
            for name in field_names:
                field = self.fields.get(name)
                if not field:
                    continue
                field.disabled = disabled
                if disabled:
                    field.widget.attrs["aria-disabled"] = "true"
                else:
                    field.widget.attrs.pop("aria-disabled", None)

        if culvert_type in {"Slab Culvert", "Box Culvert"}:
            set_disabled(self.pipe_fields, True)
            set_disabled(self.slab_box_fields, False)
        elif culvert_type == "Pipe Culvert":
            set_disabled(self.slab_box_fields, True)
            set_disabled(self.pipe_fields, False)
        else:
            set_disabled(self.pipe_fields, False)
            set_disabled(self.slab_box_fields, False)


@admin.register(models.CulvertDetail, site=grms_admin_site)
class CulvertDetailAdmin(StructureDetailOverlayMixin, RoadSectionStructureCascadeAdminMixin, GRMSBaseAdmin):
    form = CulvertDetailForm
    structure_category_codes = ("Culvert",)
    autocomplete_fields = ("structure",)
    cascade_autocomplete = {
        "structure": lambda qs, req: _filter_structure_qs(qs, req),
    }
    change_form_template = "admin/grms/structure_detail/change_form.html"
    list_display = (
        "structure",
        "culvert_type",
        "width_m",
        "span_m",
        "clear_height_m",
        "num_pipes",
        "pipe_diameter_m",
        "has_head_walls",
    )
    fieldsets = (
        ("Structure", {"fields": ("road", "section", "structure")}),
        ("Culvert type", {"fields": ("culvert_type",)}),
        (
            "Slab/Box dimensions",
            {"fields": (("width_m", "span_m", "clear_height_m"),)},
        ),
        ("Pipe details", {"fields": (("num_pipes", "pipe_diameter_m"),)}),
        ("Head walls", {"fields": ("has_head_walls",)}),
    )
    class Media:
        js = (
            "grms/admin/cascade_autocomplete.js",
            "grms/js/culvert-detail-admin.js",
        )


class FordDetailForm(StructureDetailFilterForm):
    structure_category = "Ford"

    class Meta:
        model = models.FordDetail
        fields = "__all__"


@admin.register(models.FordDetail, site=grms_admin_site)
class FordDetailAdmin(StructureDetailOverlayMixin, RoadSectionStructureCascadeAdminMixin, GRMSBaseAdmin):
    form = FordDetailForm
    structure_category_codes = ("Ford",)
    autocomplete_fields = ("structure",)
    cascade_autocomplete = {
        "structure": lambda qs, req: _filter_structure_qs(qs, req),
    }
    change_form_template = "admin/grms/structure_detail/change_form.html"
    list_display = ("structure",)
    fieldsets = (("Structure", {"fields": ("road", "section", "structure")}),)
    class Media:
        js = (
            "grms/admin/cascade_autocomplete.js",
        )


class RetainingWallDetailForm(StructureDetailFilterForm):
    structure_category = "Retaining Wall"

    class Meta:
        model = models.RetainingWallDetail
        fields = "__all__"


@admin.register(models.RetainingWallDetail, site=grms_admin_site)
class RetainingWallDetailAdmin(StructureDetailOverlayMixin, RoadSectionStructureCascadeAdminMixin, GRMSBaseAdmin):
    form = RetainingWallDetailForm
    structure_category_codes = ("Retaining Wall",)
    autocomplete_fields = ("structure",)
    cascade_autocomplete = {
        "structure": lambda qs, req: _filter_structure_qs(qs, req),
    }
    change_form_template = "admin/grms/structure_detail/change_form.html"
    list_display = ("structure",)
    fieldsets = (("Structure", {"fields": ("road", "section", "structure")}),)
    class Media:
        js = (
            "grms/admin/cascade_autocomplete.js",
        )


class GabionWallDetailForm(StructureDetailFilterForm):
    structure_category = "Gabion Wall"

    class Meta:
        model = models.GabionWallDetail
        fields = "__all__"


@admin.register(models.GabionWallDetail, site=grms_admin_site)
class GabionWallDetailAdmin(StructureDetailOverlayMixin, RoadSectionStructureCascadeAdminMixin, GRMSBaseAdmin):
    form = GabionWallDetailForm
    structure_category_codes = ("Gabion Wall",)
    autocomplete_fields = ("structure",)
    cascade_autocomplete = {
        "structure": lambda qs, req: _filter_structure_qs(qs, req),
    }
    change_form_template = "admin/grms/structure_detail/change_form.html"
    list_display = ("structure",)
    fieldsets = (("Structure", {"fields": ("road", "section", "structure")}),)
    class Media:
        js = (
            "grms/admin/cascade_autocomplete.js",
        )


@admin.register(models.FurnitureInventory, site=grms_admin_site)
class FurnitureInventoryAdmin(SectionScopedAdmin):
    class FurnitureInventoryForm(CascadeFKModelFormMixin, CascadeRoadSectionMixin, forms.ModelForm):
        road = forms.ModelChoiceField(
            queryset=models.Road.objects.all(),
            required=False,
            label="Road",
        )

        class Meta:
            model = models.FurnitureInventory
            fields = "__all__"

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            instance = self.instance
            if instance and getattr(instance, "section_id", None) and not self.is_bound:
                self.fields["road"].initial = instance.section.road
            self._setup_road_section()

        class Media:
            js = ("grms/js/furniture-admin.js",)

    form = FurnitureInventoryForm
    list_display = (
        "furniture_type",
        "section",
        "chainage_km",
        "chainage_from_km",
        "chainage_to_km",
        "left_present",
        "right_present",
    )
    list_filter = ("furniture_type",)
    search_fields = ("section__road__road_identifier", "furniture_type")
    readonly_fields = ("created_at", "modified_at")
    _AUTO = ("section",)
    autocomplete_fields = valid_autocomplete_fields(models.FurnitureInventory, _AUTO)
    fieldsets = (
        ("Furniture Info", {"fields": ("road", "section", "furniture_type")}),
        ("Point Furniture", {"fields": ("chainage_km",)}),
        ("Linear Furniture", {"fields": ("chainage_from_km", "chainage_to_km", "left_present", "right_present")}),
        ("Optional GPS", {"fields": ("location_point",)}),
        ("Comments", {"fields": ("comments",)}),
        ("Timestamps", {"fields": ("created_at", "modified_at")}),
    )

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        road_id = request.GET.get("road_id")
        if road_id and road_id.isdigit():
            queryset = queryset.filter(section__road_id=int(road_id))
        section_id = request.GET.get("section_id")
        if section_id and section_id.isdigit():
            queryset = queryset.filter(section_id=int(section_id))
        return queryset, use_distinct


class StructureConditionSurveyForm(CascadeFKModelFormMixin, CascadeRoadSectionAssetMixin, forms.ModelForm):
    road_filter = forms.ModelChoiceField(
        queryset=models.Road.objects.all(),
        required=False,
        label="Road",
    )
    section_filter = forms.ModelChoiceField(
        queryset=models.RoadSection.objects.all(),
        required=False,
        label="Section",
    )

    road_field_name = "road_filter"
    section_field_name = "section_filter"
    asset_field_name = "structure"
    asset_model = models.StructureInventory
    asset_url = "/admin/grms/options/structures/"
    asset_placeholder = "Select a structure"

    class Meta:
        model = models.StructureConditionSurvey
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        road_field = self.fields.get("road_filter")
        if road_field is not None:
            road_field.widget = AutocompleteSelect(
                models.RoadSection._meta.get_field("road"),
                grms_admin_site,
            )
        section_field = self.fields.get("section_filter")
        if section_field is not None:
            section_field.widget = AutocompleteSelect(
                models.StructureInventory._meta.get_field("section"),
                grms_admin_site,
            )
        instance = self.instance
        if instance and getattr(instance, "structure_id", None) and not self.is_bound:
            structure = instance.structure
            self.fields["road_filter"].initial = structure.road
            self.fields["section_filter"].initial = structure.section
        self._setup_road_section()
        self._setup_asset()


@admin.register(models.StructureConditionSurvey, site=grms_admin_site)
class StructureConditionSurveyAdmin(GRMSBaseAdmin):
    form = StructureConditionSurveyForm
    autocomplete_fields = ("structure", "qa_status")
    list_display = ("structure_desc", "survey_year", "condition_code", "condition_rating", "qa_status")
    list_filter = ("survey_year", "condition_rating")
    search_fields = ("structure__road__road_identifier", "structure__structure_category")
    list_select_related = ("structure", "structure__road")
    readonly_fields = ("created_at", "modified_at")
    _AUTO = ("structure", "qa_status")
    autocomplete_fields = valid_autocomplete_fields(models.StructureConditionSurvey, _AUTO)
    fieldsets = (
        ("Structure", {"fields": ("road_filter", "section_filter", "structure")}),
        (
            "Survey details",
            {
                "fields": (
                    "survey_year",
                    "condition_code",
                    "condition_rating",
                    "inspector_name",
                    "inspection_date",
                    "qa_status",
                )
            },
        ),
        ("Comments & attachments", {"fields": ("comments", "attachments")}),
        ("Audit", {"fields": ("created_at", "modified_at")}),
    )

    def structure_desc(self, obj):
        return structure_label(obj.structure)

    structure_desc.short_description = "Structure"


@admin.register(models.StructureConditionLookup, site=grms_admin_site)
class StructureConditionLookupAdmin(GRMSBaseAdmin):
    list_display = ("code", "name", "description")
    ordering = ("code",)
    search_fields = ("code", "name", "description")


@admin.register(models.StructureConditionInterventionRule, site=grms_admin_site)
class StructureConditionInterventionRuleAdmin(GRMSBaseAdmin):
    list_display = ("structure_type", "condition", "intervention_item", "is_active")
    list_filter = ("structure_type", "is_active")
    ordering = ("structure_type", "condition__code")
    autocomplete_fields = ("condition", "intervention_item")


class RoadConditionSurveyForm(CascadeFKModelFormMixin, RoadSectionSegmentFilterForm):
    class Meta:
        model = models.RoadConditionSurvey
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        road_field = self.fields.get("road")
        if road_field is not None:
            road_field.widget = AutocompleteSelect(
                models.RoadSection._meta.get_field("road"),
                grms_admin_site,
            )
        section_field = self.fields.get("section")
        if section_field is not None:
            section_field.widget = AutocompleteSelect(
                models.RoadSegment._meta.get_field("section"),
                grms_admin_site,
            )
        instance = self.instance
        if instance and getattr(instance, "road_segment_id", None):
            segment = instance.road_segment
            self.fields["road"].initial = segment.section.road
            self.fields["section"].initial = segment.section


@admin.register(models.RoadConditionSurvey, site=grms_admin_site)
class RoadConditionSurveyAdmin(
    CascadeAutocompleteAdminMixin, RoadSectionSegmentCascadeAdminMixin, SectionScopedAdmin
):
    form = RoadConditionSurveyForm
    list_display = ("road_segment", "inspection_date", "is_there_bottleneck")
    list_filter = ("inspection_date", "is_there_bottleneck")
    search_fields = ("road_segment__section__road__road_identifier", "road_segment__segment_identifier")
    autocomplete_fields = (
        "road_segment",
        "drainage_left",
        "drainage_right",
        "shoulder_left",
        "shoulder_right",
        "surface_condition",
    )
    cascade_autocomplete = {
        "road_segment": lambda qs, req: qs.filter(section_id=int(req.GET.get("section")))
        if (req.GET.get("section") or "").isdigit()
        else qs.none(),
    }
    actions = [export_condition_surveys_to_excel]

    class Media:
        js = (
            "grms/admin/cascade_autocomplete.js",
            "grms/js/cascade_static_hierarchy.js",
        )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        road_id = request.GET.get("road__id__exact")
        if road_id and road_id.isdigit():
            qs = qs.filter(road_segment__section__road_id=int(road_id))
        section_id = request.GET.get("section__id__exact")
        if section_id and section_id.isdigit():
            qs = qs.filter(road_segment__section_id=int(section_id))
        return qs

    fieldsets = (
        ("Survey Info", {
            "fields": (
                "road",
                "section",
                "road_segment",
                ("inspection_date", "inspected_by"),
            ),
            "description": "Pick the road, then section, then segment to scope the survey.",
        }),
        ("Drainage", {
            "fields": (
                ("drainage_left", "drainage_right"),
            )
        }),
        ("Shoulders", {
            "fields": (
                ("shoulder_left", "shoulder_right"),
            )
        }),
        ("Surface", {
            "fields": ("surface_condition",),
        }),
        ("Gravel Thickness", {
            "fields": ("gravel_thickness_mm",),
        }),
        ("Bottleneck", {
            "fields": ("is_there_bottleneck", "bottleneck_size_m"),
        }),
        ("Comments", {
            "fields": ("comments",),
        }),
    )


class FurnitureConditionSurveyForm(CascadeFKModelFormMixin, CascadeRoadSectionAssetMixin, forms.ModelForm):
    road_filter = forms.ModelChoiceField(
        queryset=models.Road.objects.all(),
        required=False,
        label="Road",
    )
    section_filter = forms.ModelChoiceField(
        queryset=models.RoadSection.objects.all(),
        required=False,
        label="Section",
    )

    road_field_name = "road_filter"
    section_field_name = "section_filter"
    asset_field_name = "furniture"
    asset_model = models.FurnitureInventory
    asset_url = "/admin/grms/options/furniture/"
    asset_placeholder = "Select a furniture item"

    class Meta:
        model = models.FurnitureConditionSurvey
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        road_field = self.fields.get("road_filter")
        if road_field is not None:
            road_field.widget = AutocompleteSelect(
                models.RoadSection._meta.get_field("road"),
                grms_admin_site,
            )
        section_field = self.fields.get("section_filter")
        if section_field is not None:
            section_field.widget = AutocompleteSelect(
                models.StructureInventory._meta.get_field("section"),
                grms_admin_site,
            )
        instance = self.instance
        if instance and getattr(instance, "furniture_id", None) and not self.is_bound:
            furniture = instance.furniture
            self.fields["road_filter"].initial = furniture.section.road
            self.fields["section_filter"].initial = furniture.section
        self._setup_road_section()
        self._setup_asset()


@admin.register(models.FurnitureConditionSurvey, site=grms_admin_site)
class FurnitureConditionSurveyAdmin(GRMSBaseAdmin):
    form = FurnitureConditionSurveyForm
    autocomplete_fields = ("furniture", "qa_status")
    list_display = ("furniture", "survey_year", "condition_rating")
    list_filter = ("survey_year", "condition_rating")
    readonly_fields = ("created_at",)
    _AUTO = ("furniture", "qa_status")
    autocomplete_fields = valid_autocomplete_fields(models.FurnitureConditionSurvey, _AUTO)
    fieldsets = (
        ("Furniture", {"fields": ("road_filter", "section_filter", "furniture")}),
        (
            "Survey",
            {
                "fields": (
                    "survey_year",
                    "condition_rating",
                    "qa_status",
                    "inspected_by",
                    "inspection_date",
                )
            },
        ),
        ("Notes", {"fields": ("comments",)}),
        ("Created", {"fields": ("created_at",)}),
    )


class RoadConditionDetailedSurveyForm(CascadeFKModelFormMixin, RoadSectionSegmentFilterForm):
    class Meta:
        model = models.RoadConditionDetailedSurvey
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        road_field = self.fields.get("road")
        if road_field is not None:
            road_field.widget = AutocompleteSelect(
                models.RoadSection._meta.get_field("road"),
                grms_admin_site,
            )
        section_field = self.fields.get("section")
        if section_field is not None:
            section_field.widget = AutocompleteSelect(
                models.RoadSegment._meta.get_field("section"),
                grms_admin_site,
            )
        instance = self.instance
        if instance and getattr(instance, "road_segment_id", None):
            segment = instance.road_segment
            self.fields["road"].initial = segment.section.road
            self.fields["section"].initial = segment.section


@admin.register(models.RoadConditionDetailedSurvey, site=grms_admin_site)
class RoadConditionDetailedSurveyAdmin(CascadeAutocompleteAdminMixin, SectionScopedAdmin):
    form = RoadConditionDetailedSurveyForm
    autocomplete_fields = ("awp", "road_segment", "distress", "distress_condition", "activity", "qa_status")
    list_display = ("road_segment", "distress", "survey_level", "inspection_date")
    list_filter = ("survey_level", "inspection_date", "qa_status")
    search_fields = ("road_segment__section__road__road_identifier", "distress__name")
    _AUTO = ("road_segment", "distress", "distress_condition", "activity", "qa_status", "awp")
    autocomplete_fields = valid_autocomplete_fields(models.RoadConditionDetailedSurvey, _AUTO)
    cascade_autocomplete = {
        "road_segment": lambda qs, req: qs.filter(section_id=int(req.GET.get("section")))
        if (req.GET.get("section") or "").isdigit()
        else qs.none(),
    }

    class Media:
        js = (
            "grms/admin/cascade_autocomplete.js",
            "grms/js/cascade_static_hierarchy.js",
        )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        road_id = request.GET.get("road__id__exact")
        if road_id and road_id.isdigit():
            qs = qs.filter(road_segment__section__road_id=int(road_id))
        section_id = request.GET.get("section__id__exact")
        if section_id and section_id.isdigit():
            qs = qs.filter(road_segment__section_id=int(section_id))
        return qs
    fieldsets = (
        (
            "Survey context",
            {
                "fields": (
                    "survey_level",
                    "awp",
                    "road",
                    "section",
                    "road_segment",
                    "inspected_by",
                    "inspection_date",
                    "qa_status",
                )
            },
        ),
        (
            "Distress classification",
            {
                "fields": (
                    "distress",
                    "distress_condition",
                    "severity_code",
                    "extent_code",
                    "extent_percent",
                )
            },
        ),
        (
            "Measurements",
            {
                "fields": (
                    "distress_length_m",
                    "distress_area_m2",
                    "distress_volume_m3",
                    "observed_gravel_thickness_mm",
                    "carriageway_width_m",
                    "shoulder_width_left_m",
                    "shoulder_width_right_m",
                    ("ditch_left_present", "ditch_right_present"),
                )
            },
        ),
        (
            "Activity & quantities",
            {
                "fields": (
                    "activity",
                    "quantity_unit",
                    "quantity_estimated",
                    "quantity_source",
                )
            },
        ),
        ("Notes", {"fields": ("severity_notes", "comments")}),
    )


@admin.register(models.StructureConditionDetailedSurvey, site=grms_admin_site)
class StructureConditionDetailedSurveyAdmin(GRMSBaseAdmin):
    autocomplete_fields = ("awp", "structure", "distress", "distress_condition", "activity", "qa_status")
    list_display = ("structure", "distress", "survey_level", "inspection_date")
    list_filter = ("survey_level", "inspection_date")
    _AUTO = ("structure", "distress", "distress_condition", "activity", "qa_status", "awp")
    autocomplete_fields = valid_autocomplete_fields(models.StructureConditionDetailedSurvey, _AUTO)
    fieldsets = (
        (
            "Survey context",
            {
                "fields": (
                    "survey_level",
                    "awp",
                    "structure",
                    "inspected_by",
                    "inspection_date",
                    "qa_status",
                )
            },
        ),
        (
            "Distress classification",
            {
                "fields": (
                    "distress",
                    "distress_condition",
                    "severity_code",
                    "extent_code",
                )
            },
        ),
        (
            "Measurements",
            {
                "fields": (
                    "distress_length_m",
                    "distress_area_m2",
                    "distress_volume_m3",
                    "check_dam_count",
                )
            },
        ),
        (
            "Activity & quantities",
            {
                "fields": (
                    "activity",
                    "quantity_unit",
                    "quantity_estimated",
                    "computed_by_lookup",
                )
            },
        ),
        ("Notes", {"fields": ("severity_notes", "comments")}),
    )


@admin.register(models.FurnitureConditionDetailedSurvey, site=grms_admin_site)
class FurnitureConditionDetailedSurveyAdmin(GRMSBaseAdmin):
    autocomplete_fields = ("awp", "furniture", "distress", "distress_condition", "activity", "qa_status")
    list_display = ("furniture", "distress", "survey_level", "inspection_date")
    list_filter = ("survey_level", "inspection_date")
    fieldsets = (
        (
            "Survey context",
            {
                "fields": (
                    "survey_level",
                    "awp",
                    "furniture",
                    "inspected_by",
                    "inspection_date",
                    "qa_status",
                )
            },
        ),
        (
            "Distress classification",
            {
                "fields": (
                    "distress",
                    "distress_condition",
                    "severity_code",
                    "extent_code",
                )
            },
        ),
        (
            "Activity & quantities",
            {
                "fields": (
                    "activity",
                    "quantity_unit",
                    "quantity_estimated",
                    "computed_by_lookup",
                )
            },
        ),
        ("Notes", {"fields": ("severity_notes", "comments")}),
    )


class RoadSocioEconomicForm(forms.ModelForm):
    class Meta:
        model = models.RoadSocioEconomic
        fields = "__all__"
        widgets = {"population_served": forms.NumberInput(attrs={"style": "width: 10ch;"})}


@admin.register(models.RoadSocioEconomic, site=grms_admin_site)
class RoadSocioEconomicAdmin(GRMSBaseAdmin):
    form = RoadSocioEconomicForm
    list_display = (
        "road",
        "population_served",
        "trading_centers",
        "villages",
        "markets",
        "health_centers",
        "education_centers",
    )
    list_filter = ("road__admin_zone", "road__admin_woreda")
    search_fields = ("road__road_identifier", "road__road_name_from", "road__road_name_to")
    autocomplete_fields = ("road", "road_link_type")
    fieldsets = (
        ("Context", {"fields": ("road", "population_served", "road_link_type", "notes")}),
        (
            "Transport & Connectivity (BF1)",
            {
                "fields": (
                    "adt_override",
                    "trading_centers",
                    "villages",
                )
            },
        ),
        (
            "Agriculture & Market Access (BF2)",
            {
                "fields": (
                    "farmland_percent",
                    "cooperative_centers",
                    "markets",
                )
            },
        ),
        (
            "Social Services & Development (BF3)",
            {
                "fields": (
                    "health_centers",
                    "education_centers",
                    "development_projects",
                )
            },
        ),
    )


@admin.register(models.BenefitCategory, site=grms_admin_site)
class BenefitCategoryAdmin(GRMSBaseAdmin):
    list_display = ("code", "name", "weight_display")
    search_fields = ("code", "name")

    @staticmethod
    def weight_display(obj):
        return obj.weight

    weight_display.short_description = "Weight"


@admin.register(models.BenefitCriterion, site=grms_admin_site)
class BenefitCriterionAdmin(GRMSBaseAdmin):
    list_display = ("code", "name", "category", "weight", "scoring_method")
    list_filter = ("category", "scoring_method")
    search_fields = ("code", "name")


@admin.register(models.BenefitCriterionScale, site=grms_admin_site)
class BenefitCriterionScaleAdmin(GRMSBaseAdmin):
    list_display = (
        "criterion",
        "min_value",
        "max_value",
        "score",
        "description",
    )
    list_filter = ("criterion",)
    search_fields = ("criterion__name",)


@admin.register(models.BenefitFactor, site=grms_admin_site)
class BenefitFactorAdmin(GRMSBaseAdmin):
    list_display = (
        "road",
        "fiscal_year",
        "bf1_transport_score",
        "bf2_agriculture_score",
        "bf3_social_score",
        "total_benefit_score",
    )
    list_filter = ("fiscal_year", "road__admin_zone", "road__admin_woreda")
    readonly_fields = (
        "road",
        "fiscal_year",
        "bf1_transport_score",
        "bf2_agriculture_score",
        "bf3_social_score",
        "total_benefit_score",
        "calculated_at",
        "notes",
    )
    fieldsets = (
        ("Context", {"fields": ("road", "fiscal_year", "calculated_at", "notes")}),
        (
            "Benefit factors",
            {
                "fields": (
                    "bf1_transport_score",
                    "bf2_agriculture_score",
                    "bf3_social_score",
                    "total_benefit_score",
                )
            },
        ),
    )
    _AUTO = ("road",)
    autocomplete_fields = valid_autocomplete_fields(models.BenefitFactor, _AUTO)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False if obj else super().has_change_permission(request, obj)

@admin.register(models.RoadRankingResult, site=grms_admin_site)
class RoadRankingResultAdmin(GRMSBaseAdmin):
    list_display = (
        "rank",
        "road",
        "road_class_or_surface_group",
        "road_index",
        "population_served",
        "benefit_factor",
        "cost_of_improvement",
    )
    list_filter = ("fiscal_year", "road_class_or_surface_group")
    ordering = ("rank",)
    search_fields = ("road__road_identifier", "road__road_name_from", "road__road_name_to")
    _AUTO = ("road",)
    autocomplete_fields = valid_autocomplete_fields(models.RoadRankingResult, _AUTO)
    readonly_fields = (
        "road",
        "fiscal_year",
        "road_class_or_surface_group",
        "population_served",
        "benefit_factor",
        "cost_of_improvement",
        "road_index",
        "rank",
    )

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False if obj else super().has_change_permission(request, obj)
