from __future__ import annotations

from io import BytesIO

from django.http import HttpResponse
from django.template.response import TemplateResponse
from openpyxl import Workbook

from . import reports


def _workbook_response(filename: str, workbook: Workbook) -> HttpResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def reports_index_view(request):
    context = {
        **request.admin_site.each_context(request),
        "title": "Reports",
    }
    return TemplateResponse(request, "admin/grms/reports/index.html", context)


def road_inventory_report_view(request):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Road Inventory"
    ws.append(["Road ID", "Road name", "Total length (km)", "Sections", "Segments"])
    for row in reports.road_inventory_rows():
        ws.append(
            [
                row.road_identifier,
                row.road_name,
                row.total_length_km,
                row.section_count,
                row.segment_count,
            ]
        )
    return _workbook_response("road_inventory.xlsx", workbook)


def structure_inventory_report_view(request):
    road_id = request.GET.get("road")
    road_id_int = int(road_id) if road_id and road_id.isdigit() else None

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Structure Inventory"
    ws.append(["Road ID", "Section", "Category", "Structure", "Easting (m)", "Northing (m)"])
    for row in reports.structure_inventory_rows(road_id=road_id_int):
        ws.append(
            [
                row.road_identifier,
                row.section_label,
                row.structure_category,
                row.structure_name,
                row.easting_m,
                row.northing_m,
            ]
        )
    return _workbook_response("structure_inventory.xlsx", workbook)


def condition_survey_report_view(request):
    fiscal_year = request.GET.get("fiscal_year")
    fiscal_year_int = int(fiscal_year) if fiscal_year and fiscal_year.isdigit() else None

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Condition Surveys"
    ws.append(["Road ID", "Section", "Segment", "Inspection date", "MCI"])
    for row in reports.condition_survey_rows(fiscal_year=fiscal_year_int):
        ws.append(
            [
                row.road_identifier,
                row.section_label,
                row.segment_label,
                row.inspection_date.isoformat() if row.inspection_date else "",
                row.mci_value,
            ]
        )
    return _workbook_response("condition_surveys.xlsx", workbook)
